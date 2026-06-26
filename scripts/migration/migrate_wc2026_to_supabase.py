#!/usr/bin/env python3
"""
Migrate WC2026 source data into the clean Supabase schema.

This script reads either the live source spreadsheet or an Excel workbook
export and writes only canonical/useful data into Supabase.

Required environment:
  SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY  (preferred) or SUPABASE_KEY

Supported input:
  --google-spreadsheet-id <spreadsheet_id> --google-credentials-json /path/to/service-account.json
  --xlsx /path/to/workbook.xlsx

Before running against an empty database:
  1. Run supabase/new_project/001_clean_schema.sql
  2. Run supabase/new_project/003_seed_countries_wc2026.sql
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import ssl
import time
import uuid
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, time as dt_time, timezone
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo


SOURCE_NAME = "SOURCE_SNAPSHOT"
DEFAULT_SEASON_SLUG = "wc2026"
DEFAULT_COMPETITION_SLUG = "fifa-world-cup"
DEFAULT_SOURCE_TZ = "America/Santiago"

TABLE_PRIMARY_KEYS = {
    "competitions": "competition_id",
    "competition_seasons": "competition_season_id",
    "competition_stages": "stage_id",
    "competition_groups": "group_id",
    "teams": "team_id",
    "team_aliases": "team_alias_id",
    "competition_team_entries": "competition_team_entry_id",
    "competition_group_memberships": "group_membership_id",
    "players": "player_id",
    "team_memberships": "team_membership_id",
    "competition_rosters": "competition_roster_id",
    "venues": "venue_id",
    "referees": "referee_id",
    "matches": "match_id",
    "tournament_slots": "tournament_slot_id",
    "match_participants": "match_participant_id",
    "entity_external_refs": "entity_external_ref_id",
    "match_lineups": "match_lineup_id",
    "match_events": "match_event_id",
    "match_officials": "match_official_id",
    "player_match_stats": "player_match_stat_id",
}


TEAM_DISPLAY_ES = {
    "algeria": "Argelia",
    "argentina": "Argentina",
    "australia": "Australia",
    "austria": "Austria",
    "belgium": "Belgica",
    "bosniaherzegovina": "Bosnia y Herzegovina",
    "brazil": "Brasil",
    "canada": "Canada",
    "capeverde": "Cabo Verde",
    "colombia": "Colombia",
    "congodr": "Congo DR",
    "cotedivoire": "Costa de Marfil",
    "croatia": "Croacia",
    "curacao": "Curazao",
    "czechia": "Republica Checa",
    "ecuador": "Ecuador",
    "egypt": "Egipto",
    "england": "Inglaterra",
    "france": "Francia",
    "germany": "Alemania",
    "ghana": "Ghana",
    "haiti": "Haiti",
    "iran": "Iran",
    "iraq": "Irak",
    "japan": "Japon",
    "jordan": "Jordania",
    "mexico": "Mexico",
    "morocco": "Marruecos",
    "netherlands": "Paises Bajos",
    "newzealand": "Nueva Zelanda",
    "norway": "Noruega",
    "panama": "Panama",
    "paraguay": "Paraguay",
    "portugal": "Portugal",
    "qatar": "Catar",
    "saudiarabia": "Arabia Saudita",
    "scotland": "Escocia",
    "senegal": "Senegal",
    "southafrica": "Sudafrica",
    "southkorea": "Corea del Sur",
    "spain": "Espana",
    "sweden": "Suecia",
    "switzerland": "Suiza",
    "tunisia": "Tunez",
    "turkey": "Turquia",
    "unitedstates": "EE.UU.",
    "uruguay": "Uruguay",
    "uzbekistan": "Uzbekistan",
}

TEAM_ALIASES = {
    "algeria": ["algeria", "argelia", "alg", "dz"],
    "argentina": ["argentina", "arg"],
    "australia": ["australia", "aus"],
    "austria": ["austria", "aut"],
    "belgium": ["belgium", "belgica", "bel"],
    "bosniaherzegovina": ["bosnia and herzegovina", "bosnia herzegovina", "bosnia-herzegovina", "bosnia & herzegovina", "bosnia", "bih"],
    "brazil": ["brazil", "brasil", "bra"],
    "canada": ["canada", "canada", "can"],
    "capeverde": ["cape verde", "cabo verde", "cpv"],
    "colombia": ["colombia", "col"],
    "congodr": ["dr congo", "congo dr", "drc", "cod", "rd congo"],
    "cotedivoire": ["ivory coast", "cote d ivoire", "cote d'ivoire", "costa de marfil", "civ"],
    "croatia": ["croatia", "croacia", "hrv"],
    "curacao": ["curacao", "curazao", "cuw"],
    "czechia": ["czechia", "czech republic", "republica checa", "cze"],
    "ecuador": ["ecuador", "ecu"],
    "egypt": ["egypt", "egipto", "egy"],
    "england": ["england", "inglaterra", "eng"],
    "france": ["france", "francia", "fra"],
    "germany": ["germany", "alemania", "deutschland", "ger", "deu"],
    "ghana": ["ghana", "gha"],
    "haiti": ["haiti", "hai", "hti"],
    "iran": ["iran", "irn"],
    "iraq": ["iraq", "irak", "irq"],
    "japan": ["japan", "japon", "jpn"],
    "jordan": ["jordan", "jordania", "jor"],
    "mexico": ["mexico", "mex"],
    "morocco": ["morocco", "marruecos", "mar"],
    "netherlands": ["netherlands", "holanda", "paises bajos", "ned", "nld"],
    "newzealand": ["new zealand", "nueva zelanda", "nzl"],
    "norway": ["norway", "noruega", "nor"],
    "panama": ["panama", "pan"],
    "paraguay": ["paraguay", "par"],
    "portugal": ["portugal", "por"],
    "qatar": ["qatar", "catar", "qat"],
    "saudiarabia": ["saudi arabia", "arabia saudita", "ksa", "sau"],
    "scotland": ["scotland", "escocia", "sco"],
    "senegal": ["senegal", "sen"],
    "southafrica": ["south africa", "sudafrica", "rsa", "zaf"],
    "southkorea": ["south korea", "korea republic", "republic of korea", "corea del sur", "kor"],
    "spain": ["spain", "espana", "esp"],
    "sweden": ["sweden", "suecia", "swe"],
    "switzerland": ["switzerland", "suiza", "sui", "che"],
    "tunisia": ["tunisia", "tunez", "tun"],
    "turkey": ["turkey", "turkiye", "turquia", "tur"],
    "unitedstates": ["united states", "united states of america", "usa", "u s a", "us", "eeuu", "ee uu", "estados unidos"],
    "uruguay": ["uruguay", "uru"],
    "uzbekistan": ["uzbekistan", "uzb"],
}

TEAM_COUNTRY = {
    "algeria": "DZ",
    "argentina": "AR",
    "australia": "AU",
    "austria": "AT",
    "belgium": "BE",
    "bosniaherzegovina": "BA",
    "brazil": "BR",
    "canada": "CA",
    "capeverde": "CV",
    "colombia": "CO",
    "congodr": "CD",
    "cotedivoire": "CI",
    "croatia": "HR",
    "curacao": "CW",
    "czechia": "CZ",
    "ecuador": "EC",
    "egypt": "EG",
    "england": "GB",
    "france": "FR",
    "germany": "DE",
    "ghana": "GH",
    "haiti": "HT",
    "iran": "IR",
    "iraq": "IQ",
    "japan": "JP",
    "jordan": "JO",
    "mexico": "MX",
    "morocco": "MA",
    "netherlands": "NL",
    "newzealand": "NZ",
    "norway": "NO",
    "panama": "PA",
    "paraguay": "PY",
    "portugal": "PT",
    "qatar": "QA",
    "saudiarabia": "SA",
    "scotland": "GB",
    "senegal": "SN",
    "southafrica": "ZA",
    "southkorea": "KR",
    "spain": "ES",
    "sweden": "SE",
    "switzerland": "CH",
    "tunisia": "TN",
    "turkey": "TR",
    "unitedstates": "US",
    "uruguay": "UY",
    "uzbekistan": "UZ",
}

TEAM_FIFA = {
    "algeria": "ALG",
    "argentina": "ARG",
    "australia": "AUS",
    "austria": "AUT",
    "belgium": "BEL",
    "bosniaherzegovina": "BIH",
    "brazil": "BRA",
    "canada": "CAN",
    "capeverde": "CPV",
    "colombia": "COL",
    "congodr": "COD",
    "cotedivoire": "CIV",
    "croatia": "CRO",
    "curacao": "CUW",
    "czechia": "CZE",
    "ecuador": "ECU",
    "egypt": "EGY",
    "england": "ENG",
    "france": "FRA",
    "germany": "GER",
    "ghana": "GHA",
    "haiti": "HAI",
    "iran": "IRN",
    "iraq": "IRQ",
    "japan": "JPN",
    "jordan": "JOR",
    "mexico": "MEX",
    "morocco": "MAR",
    "netherlands": "NED",
    "newzealand": "NZL",
    "norway": "NOR",
    "panama": "PAN",
    "paraguay": "PAR",
    "portugal": "POR",
    "qatar": "QAT",
    "saudiarabia": "KSA",
    "scotland": "SCO",
    "senegal": "SEN",
    "southafrica": "RSA",
    "southkorea": "KOR",
    "spain": "ESP",
    "sweden": "SWE",
    "switzerland": "SUI",
    "tunisia": "TUN",
    "turkey": "TUR",
    "unitedstates": "USA",
    "uruguay": "URU",
    "uzbekistan": "UZB",
}

VENUE_CATALOG = {
    "at-t-stadium": {
        "display_name": "AT&T Stadium",
        "aliases": ["AT&T Stadium", "AT and T Stadium", "Dallas Stadium"],
        "city": "Arlington",
        "host_city": "Dallas",
        "country_code": "US",
        "timezone_name": "America/Chicago",
        "capacity": 94000,
    },
    "bc-place": {
        "display_name": "BC Place",
        "aliases": ["BC Place", "Vancouver Stadium"],
        "city": "Vancouver",
        "host_city": "Vancouver",
        "country_code": "CA",
        "timezone_name": "America/Vancouver",
        "capacity": 54000,
    },
    "bmo-field": {
        "display_name": "BMO Field",
        "aliases": ["BMO Field", "Toronto Stadium"],
        "city": "Toronto",
        "host_city": "Toronto",
        "country_code": "CA",
        "timezone_name": "America/Toronto",
        "capacity": 45000,
    },
    "estadio-akron": {
        "display_name": "Estadio Akron",
        "aliases": ["Estadio Akron", "Akron Stadium", "Guadalajara Stadium"],
        "city": "Zapopan",
        "host_city": "Guadalajara",
        "country_code": "MX",
        "timezone_name": "America/Mexico_City",
        "capacity": 48000,
    },
    "estadio-banorte": {
        "display_name": "Estadio Banorte",
        "aliases": ["Estadio Banorte", "Estadio Azteca", "Azteca Stadium", "Mexico City Stadium", "Estadio Ciudad de Mexico"],
        "city": "Mexico City",
        "host_city": "Mexico City",
        "country_code": "MX",
        "timezone_name": "America/Mexico_City",
        "capacity": 83000,
    },
    "estadio-bbva": {
        "display_name": "Estadio BBVA",
        "aliases": ["Estadio BBVA", "BBVA Stadium", "Monterrey Stadium"],
        "city": "Guadalupe",
        "host_city": "Monterrey",
        "country_code": "MX",
        "timezone_name": "America/Monterrey",
        "capacity": 53500,
    },
    "geha-field-at-arrowhead-stadium": {
        "display_name": "GEHA Field at Arrowhead Stadium",
        "aliases": ["GEHA Field at Arrowhead Stadium", "Arrowhead Stadium", "Kansas City Stadium"],
        "city": "Kansas City",
        "host_city": "Kansas City",
        "country_code": "US",
        "timezone_name": "America/Chicago",
        "capacity": 73000,
    },
    "gillette-stadium": {
        "display_name": "Gillette Stadium",
        "aliases": ["Gillette Stadium", "Boston Stadium"],
        "city": "Foxborough",
        "host_city": "Boston",
        "country_code": "US",
        "timezone_name": "America/New_York",
        "capacity": 65000,
    },
    "hard-rock-stadium": {
        "display_name": "Hard Rock Stadium",
        "aliases": ["Hard Rock Stadium", "Miami Stadium"],
        "city": "Miami Gardens",
        "host_city": "Miami",
        "country_code": "US",
        "timezone_name": "America/New_York",
        "capacity": 65000,
    },
    "levis-stadium": {
        "display_name": "Levi's Stadium",
        "aliases": ["Levi's Stadium", "Levis Stadium", "San Francisco Bay Area Stadium"],
        "city": "Santa Clara",
        "host_city": "San Francisco Bay Area",
        "country_code": "US",
        "timezone_name": "America/Los_Angeles",
        "capacity": 71000,
    },
    "lincoln-financial-field": {
        "display_name": "Lincoln Financial Field",
        "aliases": ["Lincoln Financial Field", "Philadelphia Stadium"],
        "city": "Philadelphia",
        "host_city": "Philadelphia",
        "country_code": "US",
        "timezone_name": "America/New_York",
        "capacity": 69000,
    },
    "lumen-field": {
        "display_name": "Lumen Field",
        "aliases": ["Lumen Field", "Seattle Stadium"],
        "city": "Seattle",
        "host_city": "Seattle",
        "country_code": "US",
        "timezone_name": "America/Los_Angeles",
        "capacity": 69000,
    },
    "mercedes-benz-stadium": {
        "display_name": "Mercedes-Benz Stadium",
        "aliases": ["Mercedes-Benz Stadium", "Mercedes Benz Stadium", "Atlanta Stadium"],
        "city": "Atlanta",
        "host_city": "Atlanta",
        "country_code": "US",
        "timezone_name": "America/New_York",
        "capacity": 75000,
    },
    "metlife-stadium": {
        "display_name": "MetLife Stadium",
        "aliases": ["MetLife Stadium", "New York New Jersey Stadium"],
        "city": "East Rutherford",
        "host_city": "New York/New Jersey",
        "country_code": "US",
        "timezone_name": "America/New_York",
        "capacity": 82500,
    },
    "nrg-stadium": {
        "display_name": "NRG Stadium",
        "aliases": ["NRG Stadium", "Houston Stadium"],
        "city": "Houston",
        "host_city": "Houston",
        "country_code": "US",
        "timezone_name": "America/Chicago",
        "capacity": 72220,
    },
    "sofi-stadium": {
        "display_name": "SoFi Stadium",
        "aliases": ["SoFi Stadium", "Sofi Stadium", "Los Angeles Stadium"],
        "city": "Inglewood",
        "host_city": "Los Angeles",
        "country_code": "US",
        "timezone_name": "America/Los_Angeles",
        "capacity": 70000,
    },
}


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


ALIAS_TO_KEY = {
    normalize_text(alias).replace(" ", ""): key
    for key, aliases in TEAM_ALIASES.items()
    for alias in aliases
}


def team_key(name: Any) -> str:
    normalized = normalize_text(name)
    compact = normalized.replace(" ", "")
    return ALIAS_TO_KEY.get(compact, compact)


def display_name_for_team(name: Any) -> str:
    key = team_key(name)
    return TEAM_DISPLAY_ES.get(key, str(name or "").strip())


def slugify(value: Any) -> str:
    return normalize_text(value).replace(" ", "-")


VENUE_ALIAS_TO_SLUG = {
    slugify(alias): slug
    for slug, venue in VENUE_CATALOG.items()
    for alias in venue["aliases"]
}


def is_tournament_slot(name: Any) -> bool:
    s = normalize_text(name)
    if not s:
        return False
    patterns = [
        r"^group [a-z0-9]+ (winner|runner up|2nd place|second place|third place|3rd place)$",
        r"^third place group",
        r"^round of [0-9]+ [0-9]+ (winner|loser)$",
        r"^quarter ?final [0-9]+ (winner|loser)$",
        r"^semi ?final [0-9]+ (winner|loser)$",
        r"^semifinal [0-9]+ (winner|loser)$",
    ]
    return any(re.search(pattern, s) for pattern in patterns)


def slot_code(name: Any) -> str:
    return normalize_text(name).replace(" ", "_")


def first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip() != "":
            return value
    return None


def to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace("%", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def source_id_text(value: Any) -> str | None:
    if value is None or value == "":
        return None
    as_int = to_int(value)
    if as_int is not None:
        return str(as_int)
    return str(value).strip()


def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date, dt_time)):
        return value.isoformat()
    return str(value)


def parse_datetime(row: dict[str, Any], source_tz: str) -> str | None:
    for key in ("date_utc", "fecha_utc", "kickoff_at"):
        value = row.get(key)
        if value:
            dt = coerce_datetime(value, ZoneInfo("UTC"))
            if dt:
                return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")

    fecha = first_present(row, "fecha", "fecha_chile", "date")
    hora = first_present(row, "hora_chile", "hora", "time")
    if not fecha:
        return None

    local_tz = ZoneInfo(source_tz)
    if isinstance(fecha, datetime):
        d = fecha.date()
    elif isinstance(fecha, date):
        d = fecha
    else:
        text = str(fecha).strip()
        d = None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                d = datetime.strptime(text[:10], fmt).date()
                break
            except ValueError:
                pass
        if d is None:
            dt = coerce_datetime(text, local_tz)
            if dt:
                return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
            return None

    if isinstance(hora, datetime):
        t = hora.time()
    elif isinstance(hora, dt_time):
        t = hora
    elif hora:
        match = re.search(r"(\d{1,2}):(\d{2})", str(hora))
        t = dt_time(int(match.group(1)), int(match.group(2))) if match else dt_time(0, 0)
    else:
        t = dt_time(0, 0)

    return datetime.combine(d, t, local_tz).astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def coerce_datetime(value: Any, default_tz: ZoneInfo) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=default_tz)
    if isinstance(value, date):
        return datetime.combine(value, dt_time(0, 0), default_tz)
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=default_tz)
    except ValueError:
        return None


@dataclass
class Supabase:
    url: str
    key: str
    dry_run: bool = False
    sleep_seconds: float = 0.0
    verify_tls: bool = True

    @property
    def rest_url(self) -> str:
        return self.url.rstrip("/") + "/rest/v1"

    def request(self, method: str, path: str, body: Any = None, query: dict[str, str] | None = None, prefer: str | None = None) -> Any:
        query_string = urllib.parse.urlencode(query or {})
        url = f"{self.rest_url}/{path}" + (f"?{query_string}" if query_string else "")
        payload = None if body is None else json.dumps(body, ensure_ascii=False, default=json_default).encode("utf-8")
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        if self.dry_run and method.upper() in {"POST", "PATCH", "DELETE"}:
            print(f"[dry-run] {method.upper()} {path} rows={len(body) if isinstance(body, list) else 1}")
            if prefer and "return=representation" in prefer:
                rows = body if isinstance(body, list) else [body]
                return [self._dry_row(path, row) for row in rows]
            return []
        req = urllib.request.Request(url, data=payload, headers=headers, method=method.upper())
        for attempt in range(4):
            try:
                with urllib.request.urlopen(req, timeout=60, context=build_ssl_context(self.verify_tls)) as response:
                    text = response.read().decode("utf-8")
                    if self.sleep_seconds:
                        time.sleep(self.sleep_seconds)
                    return json.loads(text) if text else []
            except urllib.error.HTTPError as exc:
                text = exc.read().decode("utf-8")
                if exc.code in {429, 500, 502, 503, 504} and attempt < 3:
                    time.sleep(2 ** attempt)
                    continue
                raise RuntimeError(f"Supabase HTTP {exc.code} {method.upper()} {path}: {text}") from exc
            except urllib.error.URLError as exc:
                if "CERTIFICATE_VERIFY_FAILED" in str(exc.reason):
                    raise RuntimeError(
                        "TLS certificate verification failed while connecting to Supabase. "
                        "Install/update local CA certificates or install certifi with: "
                        "python3 -m pip install certifi. As a last resort, rerun with "
                        "--insecure-skip-tls-verify."
                    ) from exc
                raise

    def _dry_row(self, path: str, row: dict[str, Any]) -> dict[str, Any]:
        table = path.split("?")[0]
        out = dict(row)
        primary_key = TABLE_PRIMARY_KEYS.get(table)
        if primary_key and primary_key not in out:
            out[primary_key] = str(uuid.uuid5(uuid.NAMESPACE_URL, table + ":" + json.dumps(row, sort_keys=True, default=json_default)))
        return out

    def upsert(self, table: str, rows: list[dict[str, Any]], conflict: str, returning: bool = False) -> list[dict[str, Any]]:
        if not rows:
            return []
        deduped = dedupe(rows, conflict.split(","))
        prefer = "resolution=merge-duplicates,return=representation" if returning else "resolution=merge-duplicates,return=minimal"
        out: list[dict[str, Any]] = []
        for batch in chunks(deduped, 500):
            result = self.request("POST", table, batch, {"on_conflict": conflict}, prefer)
            if returning and isinstance(result, list):
                out.extend(result)
        return out

    def select(self, table: str, query: dict[str, str]) -> list[dict[str, Any]]:
        if self.dry_run and table == "countries":
            raw = str(query.get("code_alpha2", ""))
            codes = raw.removeprefix("in.(").removesuffix(")").split(",") if raw.startswith("in.(") else []
            return [{"code_alpha2": code} for code in codes if code]
        if self.dry_run and table == "competition_groups":
            season_id = str(query.get("competition_season_id", "")).replace("eq.", "")
            return [{
                "group_id": str(uuid.uuid5(uuid.NAMESPACE_URL, f"competition_groups:{season_id}:Grupo {letter}")),
                "competition_season_id": season_id,
                "stage_id": str(uuid.uuid5(uuid.NAMESPACE_URL, f"stage:{season_id}:GROUP_STAGE")),
                "group_code": f"Grupo {letter}",
                "group_name": f"Grupo {letter}",
                "group_order": index + 1,
            } for index, letter in enumerate("ABCDEFGHIJKL")]
        return self.request("GET", table, None, query)

    def update(self, table: str, row: dict[str, Any], query: dict[str, str]) -> None:
        if self.dry_run:
            print(json.dumps({"dry_run_update": table, "query": query, "row": row}, ensure_ascii=False, default=json_default))
            return
        self.request("PATCH", table, row, query, "return=minimal")


@dataclass
class BudgetedApiClient:
    source: str
    base_url: str
    headers: dict[str, str]
    budget: int = 0
    cache_dir: str | None = None
    verify_tls: bool = True
    used: int = 0
    fail_on_error: bool = True

    def get(self, path: str, query: dict[str, Any] | None = None) -> dict[str, Any] | list[Any] | None:
        if self.budget <= self.used:
            return None
        query_string = urllib.parse.urlencode({k: v for k, v in (query or {}).items() if v is not None})
        url = self.base_url.rstrip("/") + "/" + path.lstrip("/") + (f"?{query_string}" if query_string else "")
        cache_path = self._cache_path(url)
        if cache_path and cache_path.exists():
            return json.loads(cache_path.read_text(encoding="utf-8"))

        headers = {
            "Accept": "application/json",
            "User-Agent": "PoolTeam2026-Migration/1.0 (+https://localhost)",
            **self.headers,
        }
        req = urllib.request.Request(url, headers=headers, method="GET")
        started = time.time()
        try:
            with urllib.request.urlopen(req, timeout=60, context=build_ssl_context(self.verify_tls)) as response:
                payload = response.read().decode("utf-8")
                self.used += 1
                data = json.loads(payload) if payload else {}
                if cache_path:
                    cache_path.parent.mkdir(parents=True, exist_ok=True)
                    cache_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
                print(json.dumps({
                    "source": self.source,
                    "endpoint": path,
                    "status": response.status,
                    "latency_ms": int((time.time() - started) * 1000),
                    "budget_used": self.used,
                    "budget_limit": self.budget,
                }, ensure_ascii=False))
                return data
        except urllib.error.HTTPError as exc:
            text = exc.read().decode("utf-8")
            self.used += 1
            print(json.dumps({
                "source": self.source,
                "endpoint": path,
                "status": exc.code,
                "latency_ms": int((time.time() - started) * 1000),
                "budget_used": self.used,
                "budget_limit": self.budget,
                "error": text[:500],
            }, ensure_ascii=False))
            if not self.fail_on_error:
                return None
            raise RuntimeError(f"{self.source} HTTP {exc.code} GET {path}: {text}") from exc

    def _cache_path(self, url: str) -> Path | None:
        if not self.cache_dir:
            return None
        digest = hashlib.sha256(url.encode("utf-8")).hexdigest()
        return Path(self.cache_dir) / self.source.lower() / f"{digest}.json"


def dedupe(rows: list[dict[str, Any]], keys: list[str]) -> list[dict[str, Any]]:
    seen: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        seen[tuple(row.get(k) for k in keys)] = row
    return list(seen.values())


def chunks(items: list[dict[str, Any]], size: int) -> Iterable[list[dict[str, Any]]]:
    for i in range(0, len(items), size):
        yield items[i : i + size]


def build_ssl_context(verify_tls: bool) -> ssl.SSLContext:
    if not verify_tls:
        return ssl._create_unverified_context()
    try:
        import certifi

        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        return ssl.create_default_context()


def load_xlsx(path: str) -> dict[str, list[dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required for --xlsx. Install it or use the bundled Python environment.") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            data[sheet_name] = []
            continue
        headers = [normalize_header(value) for value in rows[0]]
        records = []
        for values in rows[1:]:
            record = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                record[header] = values[index] if index < len(values) else None
            if any(value not in (None, "") for value in record.values()):
                records.append(record)
        data[sheet_name] = records
    return data


def load_google_spreadsheet(spreadsheet_id: str, credentials_json: str | None) -> dict[str, list[dict[str, Any]]]:
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as exc:
        raise SystemExit(
            "gspread and google-auth are required for --google-spreadsheet-id. "
            "Install them in the Python environment used by this script."
        ) from exc

    credentials_path = credentials_json or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not credentials_path:
        raise SystemExit("Set --google-credentials-json or GOOGLE_APPLICATION_CREDENTIALS.")

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    credentials = Credentials.from_service_account_file(credentials_path, scopes=scopes)
    client = gspread.authorize(credentials)
    spreadsheet = client.open_by_key(spreadsheet_id)

    data: dict[str, list[dict[str, Any]]] = {}
    for worksheet in spreadsheet.worksheets():
        values = worksheet.get_all_values()
        if not values:
            data[worksheet.title] = []
            continue
        headers = [normalize_header(value) for value in values[0]]
        records = []
        for values_row in values[1:]:
            record = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                record[header] = values_row[index] if index < len(values_row) else None
            if any(value not in (None, "") for value in record.values()):
                records.append(record)
        data[worksheet.title] = records
    return data


def normalize_header(value: Any) -> str:
    return normalize_text(value).replace(" ", "_")


def get_rows(data: dict[str, list[dict[str, Any]]], sheet_name: str) -> list[dict[str, Any]]:
    return data.get(sheet_name, [])


def load_venue_coordinates(path: str | None) -> dict[str, dict[str, float]]:
    if not path:
        return {}
    venue_file = Path(path).expanduser()
    if not venue_file.exists():
        raise SystemExit(f"Venue coordinates file not found: {venue_file}")
    lines = venue_file.read_text(encoding="utf-8").splitlines()
    if not lines:
        return {}
    coordinates: dict[str, dict[str, float]] = {}
    for line in lines[1:]:
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        name, latitude, longitude = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if not name:
            continue
        coordinates[canonical_venue_slug(name)] = {
            "latitude": float(latitude),
            "longitude": float(longitude),
        }
    return coordinates


def assert_required_countries_exist(sb: Supabase) -> None:
    required = sorted(set(TEAM_COUNTRY.values()) | {venue["country_code"] for venue in VENUE_CATALOG.values()})
    rows = sb.select("countries", {
        "select": "code_alpha2",
        "code_alpha2": f"in.({','.join(required)})",
    })
    existing = {row.get("code_alpha2") for row in rows}
    missing = [code for code in required if code not in existing]
    if missing:
        raise SystemExit(
            "Missing required countries in Supabase: "
            + ", ".join(missing)
            + ". Run supabase/new_project/003_seed_countries_wc2026.sql before the migration. "
            + "If you used 002_truncate_all_data.sql previously, rerun the countries seed because older versions truncated countries."
        )


def migrate(data: dict[str, list[dict[str, Any]]], sb: Supabase, args: argparse.Namespace) -> None:
    assert_required_countries_exist(sb)
    print("Ensuring competition catalog")
    competition = sb.upsert(
        "competitions",
        [{
            "slug": args.competition_slug,
            "display_name": "FIFA World Cup",
            "competition_type": "TOURNAMENT",
            "country_code": None,
            "region": "Global",
            "tier": 1,
            "is_international": True,
            "metadata": {"source": SOURCE_NAME, "format": "GROUPS_THEN_KNOCKOUT"},
        }],
        "slug",
        returning=True,
    )[0]
    season = sb.upsert(
        "competition_seasons",
        [{
            "competition_id": competition["competition_id"],
            "slug": args.season_slug,
            "season_label": "2026",
            "starts_at": "2026-06-11T00:00:00Z",
            "ends_at": "2026-07-19T23:59:59Z",
            "timezone_name": "UTC",
            "status": "ACTIVE",
            "format_code": "GROUPS_THEN_KNOCKOUT",
            "metadata": {"source": SOURCE_NAME, "host_countries": ["US", "MX", "CA"]},
        }],
        "slug",
        returning=True,
    )[0]
    season_id = season["competition_season_id"]

    stages = upsert_stages_and_groups(sb, season_id)
    teams = upsert_teams(sb, data, season_id)
    upsert_standings_from_classification(sb, data, season_id, stages, teams)
    players = upsert_players_and_rosters(sb, data, season_id, teams)
    venue_coordinates = load_venue_coordinates(args.venues_file)
    venues = upsert_venues(sb, data, venue_coordinates)
    matches = upsert_matches(sb, data, season_id, stages, teams, venues, args.source_timezone)
    upsert_external_refs_from_source(sb, data, teams, matches, venues)
    upsert_match_detail_sheets(sb, data, season_id, matches, teams, players)
    enrich_from_optional_apis(sb, args, teams, players, matches, venues)

    counts = {
        "teams": len(teams),
        "players": len(players),
        "venues": len(venues),
    }
    print(json.dumps({"ok": True, "counts": counts}, indent=2, ensure_ascii=False))


def upsert_stages_and_groups(sb: Supabase, season_id: str) -> dict[str, dict[str, Any]]:
    stage_rows = [
        ("GROUP_STAGE", "Fase de grupos", 1, "GROUP_STAGE", {"view_type": "GROUP_TABLES", "expected_matches": 72, "teams_per_group": 4, "group_count": 12, "qualifies": {"top_n_per_group": 2, "best_third_places": 8}, "tie_breakers": ["points", "goal_difference", "goals_for", "head_to_head"]}),
        ("ROUND_OF_32", "Dieciseisavos de final", 2, "KNOCKOUT", {"view_type": "BRACKET_ROUND", "expected_matches": 16, "single_leg": True, "extra_time": True, "penalties": True}),
        ("ROUND_OF_16", "Octavos de final", 3, "KNOCKOUT", {"view_type": "BRACKET_ROUND", "expected_matches": 8, "single_leg": True, "extra_time": True, "penalties": True}),
        ("QUARTER_FINAL", "Cuartos de final", 4, "KNOCKOUT", {"view_type": "BRACKET_ROUND", "expected_matches": 4, "single_leg": True, "extra_time": True, "penalties": True}),
        ("SEMI_FINAL", "Semifinal", 5, "KNOCKOUT", {"view_type": "BRACKET_ROUND", "expected_matches": 2, "single_leg": True, "extra_time": True, "penalties": True}),
        ("THIRD_PLACE", "Tercer lugar", 6, "THIRD_PLACE", {"view_type": "BRACKET_ROUND", "expected_matches": 1, "single_leg": True, "extra_time": True, "penalties": True}),
        ("FINAL", "Final", 7, "FINAL", {"view_type": "BRACKET_ROUND", "expected_matches": 1, "single_leg": True, "extra_time": True, "penalties": True}),
    ]
    stages = sb.upsert(
        "competition_stages",
        [{
            "competition_season_id": season_id,
            "stage_code": code,
            "stage_name": name,
            "stage_order": order,
            "stage_type": stage_type,
            "rules": rules | {"source": SOURCE_NAME},
        } for code, name, order, stage_type, rules in stage_rows],
        "competition_season_id,stage_code",
        returning=True,
    )
    by_code = {row["stage_code"]: row for row in stages}
    group_stage_id = by_code["GROUP_STAGE"]["stage_id"]
    sb.upsert(
        "competition_groups",
        [{
            "competition_season_id": season_id,
            "stage_id": group_stage_id,
            "group_code": f"Grupo {letter}",
            "group_name": f"Grupo {letter}",
            "group_order": index + 1,
            "metadata": {"source": SOURCE_NAME},
        } for index, letter in enumerate("ABCDEFGHIJKL")],
        "competition_season_id,stage_id,group_code",
    )
    return by_code


def team_group_code(team_key_value: str, data: dict[str, list[dict[str, Any]]]) -> str | None:
    for row in get_rows(data, "Clasificacion"):
        if team_key(first_present(row, "equipo")) == team_key_value:
            group = str(first_present(row, "grupo") or "").strip()
            return group or None
    return None


def upsert_teams(sb: Supabase, data: dict[str, list[dict[str, Any]]], season_id: str) -> dict[str, dict[str, Any]]:
    candidates: dict[str, dict[str, Any]] = {}
    for row in get_rows(data, "Clasificacion"):
        name = first_present(row, "equipo")
        if name:
            candidates[team_key(name)] = {"name": name, "group": first_present(row, "grupo"), "source_id": first_present(row, "equipo_id")}
    if not candidates:
        for sheet in ("Equipos", "Partidos", "SourceFixtures", "Planteles"):
            for row in get_rows(data, sheet):
                for name in candidate_team_names(sheet, row):
                    if name and not is_tournament_slot(name):
                        candidates.setdefault(team_key(name), {"name": name, "group": first_present(row, "grupo", "group_name"), "source_id": None})

    rows = []
    for key, item in sorted(candidates.items()):
        display = TEAM_DISPLAY_ES.get(key, display_name_for_team(item["name"]))
        rows.append({
            "slug": slugify(display),
            "team_type": "NATIONAL_TEAM",
            "display_name": display,
            "normalized_name": normalize_text(display),
            "country_code": TEAM_COUNTRY.get(key),
            "gender": "MEN",
            "metadata": {
                "source": SOURCE_NAME,
                "canonical_key": key,
                "names": {"es": display, "en": english_name_for_key(key)},
                "sports_entity": {"fifa_code": TEAM_FIFA.get(key)},
            },
        })
    missing_country = [row["display_name"] for row in rows if not row["country_code"]]
    if missing_country:
        raise SystemExit(f"Teams without ISO country mapping: {missing_country}")
    saved = sb.upsert("teams", rows, "slug", returning=True)
    teams = {row["metadata"]["canonical_key"]: row for row in saved}

    alias_rows = []
    ref_rows = []
    entry_rows = []
    for key, team in teams.items():
        aliases = set(TEAM_ALIASES.get(key, []))
        aliases.add(team["display_name"])
        aliases.add(english_name_for_key(key))
        for alias in aliases:
            if alias:
                alias_rows.append({
                    "team_id": team["team_id"],
                    "alias": alias,
                    "normalized_alias": normalize_text(alias),
                    "language_code": "es" if alias == team["display_name"] else None,
                    "source": SOURCE_NAME,
                    "confidence": 1,
                })
        if TEAM_FIFA.get(key):
            ref_rows.append({
                "entity_type": "TEAM",
                "entity_id": team["team_id"],
                "source": "FIFA",
                "source_entity_type": "association",
                "source_entity_id": TEAM_FIFA[key],
                "source_entity_name": team["display_name"],
                "confidence": 1,
                "is_primary": True,
                "payload": {"source": SOURCE_NAME},
            })
        entry_rows.append({
            "competition_season_id": season_id,
            "team_id": team["team_id"],
            "entry_status": "ACTIVE",
            "metadata": {"source": SOURCE_NAME},
        })
    sb.upsert("team_aliases", alias_rows, "normalized_alias,source")
    sb.upsert("entity_external_refs", ref_rows, "entity_type,source,source_entity_id")
    entries = sb.upsert("competition_team_entries", entry_rows, "competition_season_id,team_id", returning=True)
    entry_by_team = {row["team_id"]: row for row in entries}

    group_rows = sb.select("competition_groups", {"select": "*", "competition_season_id": f"eq.{season_id}"})
    groups = {row["group_code"]: row for row in group_rows}
    membership_rows = []
    for row in get_rows(data, "Clasificacion"):
        key = team_key(first_present(row, "equipo"))
        team = teams.get(key)
        group = groups.get(str(first_present(row, "grupo") or "").strip())
        if team and group:
            membership_rows.append({
                "group_id": group["group_id"],
                "competition_team_entry_id": entry_by_team[team["team_id"]]["competition_team_entry_id"],
                "membership_status": "ACTIVE",
                "seed_position": to_int(first_present(row, "posicion")),
                "metadata": {"source": SOURCE_NAME},
            })
    sb.upsert("competition_group_memberships", membership_rows, "group_id,competition_team_entry_id")

    if len(teams) != 48:
        raise SystemExit(f"Expected 48 teams, got {len(teams)}. Refusing to continue.")
    return teams


def candidate_team_names(sheet: str, row: dict[str, Any]) -> list[Any]:
    if sheet == "Equipos":
        return [first_present(row, "nombre", "equipo", "team", "display_name")]
    if sheet == "Partidos":
        return [first_present(row, "local", "home_team_name"), first_present(row, "visitante", "away_team_name")]
    if sheet == "SourceFixtures":
        return [first_present(row, "home_team_name"), first_present(row, "away_team_name")]
    if sheet == "Planteles":
        return [first_present(row, "equipo")]
    return []


def upsert_standings_from_classification(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    season_id: str,
    stages: dict[str, dict[str, Any]],
    teams: dict[str, dict[str, Any]],
) -> None:
    group_rows = sb.select("competition_groups", {"select": "*", "competition_season_id": f"eq.{season_id}"})
    groups = {row["group_code"]: row for row in group_rows}
    group_stage = stages.get("GROUP_STAGE") or {}
    rows = []
    for row in get_rows(data, "Clasificacion"):
        team = teams.get(team_key(first_present(row, "equipo")))
        group = groups.get(str(first_present(row, "grupo") or "").strip())
        if not team:
            continue
        as_of = parse_datetime({"date_utc": first_present(row, "updated_at")}, "UTC") or datetime.now(timezone.utc).isoformat()
        rows.append({
            "competition_season_id": season_id,
            "stage_id": group_stage.get("stage_id"),
            "group_id": group["group_id"] if group else None,
            "team_id": team["team_id"],
            "position": to_int(first_present(row, "posicion")),
            "played": to_int(first_present(row, "pj")),
            "wins": to_int(first_present(row, "pg")),
            "draws": to_int(first_present(row, "pe")),
            "losses": to_int(first_present(row, "pp")),
            "goals_for": to_int(first_present(row, "gf")),
            "goals_against": to_int(first_present(row, "gc")),
            "goal_difference": to_int(first_present(row, "gd")),
            "points": to_int(first_present(row, "puntos")),
            "as_of": as_of,
            "source": SOURCE_NAME,
            "payload": compact_payload(row),
        })
    sb.upsert("standings", rows, "competition_season_id,stage_id,group_id,team_id,source")


def english_name_for_key(key: str) -> str:
    aliases = TEAM_ALIASES.get(key) or [key]
    return aliases[0].title().replace(" And ", " and ")


def upsert_players_and_rosters(sb: Supabase, data: dict[str, list[dict[str, Any]]], season_id: str, teams: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    player_rows = []
    roster_candidates = []
    for row in get_rows(data, "Planteles"):
        player_name = first_present(row, "jugador", "player_name", "nombre")
        team_name = first_present(row, "equipo", "team")
        if not player_name or not team_name:
            continue
        key = team_key(team_name)
        team = teams.get(key)
        if not team:
            continue
        player_id_source = source_id_text(first_present(row, "player_id", "jugador_id"))
        slug = slugify(player_name) if not player_id_source else f"player-api-football-{player_id_source}"
        player_rows.append({
            "slug": slug,
            "display_name": str(player_name).strip(),
            "normalized_name": normalize_text(player_name),
            "nationality_country_code": team["country_code"],
            "gender": "MEN",
            "metadata": {"source": SOURCE_NAME},
        })
        roster_candidates.append((slug, team["team_id"], row))

    saved = sb.upsert("players", player_rows, "slug", returning=True)
    players = {row["slug"]: row for row in saved}
    roster_rows = []
    membership_rows = []
    alias_rows = []
    ref_rows = []
    for slug, team_id, source_row in roster_candidates:
        player = players.get(slug)
        if not player:
            continue
        player["_team_id"] = team_id
        player["_lookup_name"] = normalize_text(player["display_name"])
        player_source_id = source_id_text(first_present(source_row, "player_id", "jugador_id"))
        source_team_id = source_id_text(first_present(source_row, "team_id", "equipo_id"))
        squad_id = source_id_text(first_present(source_row, "squad_id"))
        source_text = str(first_present(source_row, "fuente") or "")
        source_name = "API_FOOTBALL" if "API_FOOTBALL" in source_text.upper() or player_source_id else SOURCE_NAME
        alias_rows.append({
            "player_id": player["player_id"],
            "alias": player["display_name"],
            "normalized_alias": normalize_text(player["display_name"]),
            "language_code": None,
            "source": f"{source_name}:TEAM:{team_id}",
            "confidence": 1,
        })
        if player_source_id:
            ref_rows.append({
                "entity_type": "PLAYER",
                "entity_id": player["player_id"],
                "source": "API_FOOTBALL",
                "source_entity_type": "player",
                "source_entity_id": player_source_id,
                "source_entity_name": player["display_name"],
                "confidence": 1,
                "is_primary": True,
                "payload": compact_payload(source_row),
            })
        if squad_id:
            ref_rows.append({
                "entity_type": "PLAYER",
                "entity_id": player["player_id"],
                "source": "API_FOOTBALL",
                "source_entity_type": "squad_entry",
                "source_entity_id": squad_id,
                "source_entity_name": player["display_name"],
                "confidence": 1,
                "is_primary": False,
                "payload": compact_payload(source_row),
            })
        if source_team_id:
            ref_rows.append({
                "entity_type": "TEAM",
                "entity_id": team_id,
                "source": "API_FOOTBALL",
                "source_entity_type": "team",
                "source_entity_id": source_team_id,
                "source_entity_name": str(first_present(source_row, "equipo", "team") or ""),
                "confidence": 0.95,
                "is_primary": False,
                "payload": {"source": SOURCE_NAME, "from": "Planteles"},
            })
        membership_rows.append({
            "player_id": player["player_id"],
            "team_id": team_id,
            "membership_type": "NATIONAL_TEAM",
            "source": SOURCE_NAME,
            "confidence": 1,
            "metadata": {},
        })
        roster_rows.append({
            "competition_season_id": season_id,
            "team_id": team_id,
            "player_id": player["player_id"],
            "shirt_number": to_int(first_present(source_row, "numero", "number")),
            "position": first_present(source_row, "posicion", "position"),
            "roster_status": "ACTIVE",
            "metadata": {"source": SOURCE_NAME, "role": first_present(source_row, "rol")},
        })
    sb.upsert("team_memberships", membership_rows, "player_id,team_id,membership_type,source")
    sb.upsert("competition_rosters", roster_rows, "competition_season_id,team_id,player_id")
    sb.upsert("player_aliases", alias_rows, "normalized_alias,source")
    sb.upsert("entity_external_refs", ref_rows, "entity_type,source,source_entity_id")
    return players


def upsert_venues(sb: Supabase, data: dict[str, list[dict[str, Any]]], coordinates: dict[str, dict[str, float]]) -> dict[str, dict[str, Any]]:
    source_venue_names: dict[str, set[str]] = {}
    for row in get_rows(data, "Partidos"):
        name = first_present(row, "estadio", "venue_name")
        if not name:
            continue
        source_venue_names.setdefault(canonical_venue_slug(name), set()).add(str(name).strip())

    unknown = sorted(slug for slug in source_venue_names if slug not in VENUE_CATALOG)
    if unknown:
        raise SystemExit(f"Unknown venues in source data: {unknown}")

    rows = []
    for slug, venue in VENUE_CATALOG.items():
        aliases = sorted(set(venue["aliases"]) | source_venue_names.get(slug, set()))
        coordinate = coordinates.get(slug, {})
        rows.append({
            "slug": slug,
            "display_name": venue["display_name"],
            "city": venue["city"],
            "country_code": venue["country_code"],
            "timezone_name": venue["timezone_name"],
            "latitude": coordinate.get("latitude"),
            "longitude": coordinate.get("longitude"),
            "metadata": {
                "source": SOURCE_NAME,
                "host_city": venue["host_city"],
                "capacity": venue["capacity"],
                "aliases": aliases,
                "fifa_world_cup_2026": True,
                "coordinates_source": "VENUES_FILE" if coordinate else None,
            },
        })
    saved = sb.upsert("venues", rows, "slug", returning=True)
    if len(saved) != 16:
        raise SystemExit(f"Expected 16 canonical venues, got {len(saved)}. Refusing to continue.")
    return {row["slug"]: row for row in saved}


def canonical_venue_slug(value: Any) -> str:
    return VENUE_ALIAS_TO_SLUG.get(slugify(value), slugify(value))


def upsert_external_refs_from_source(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    teams: dict[str, dict[str, Any]],
    matches: dict[str, dict[str, Any]],
    venues: dict[str, dict[str, Any]],
) -> None:
    ref_rows = []

    for venue in venues.values():
        ref_rows.append({
            "entity_type": "VENUE",
            "entity_id": venue["venue_id"],
            "source": "FIFA_WORLD_CUP_2026_VENUE_CATALOG",
            "source_entity_type": "venue",
            "source_entity_id": venue["slug"],
            "source_entity_name": venue["display_name"],
            "confidence": 1,
            "is_primary": True,
            "payload": venue.get("metadata") or {},
        })

    for row in get_rows(data, "SourceFixtures"):
        source = str(first_present(row, "source") or "").strip().upper()
        source_match_id = first_present(row, "source_match_id", "fixture_id", "id")
        source_fixture_key = first_present(row, "source_fixture_key")
        home_name = first_present(row, "home_team_name")
        away_name = first_present(row, "away_team_name")
        kickoff = parse_datetime(row, DEFAULT_SOURCE_TZ)
        match_slug = slugify(source_fixture_key) if source_fixture_key else None
        if not match_slug and kickoff and home_name and away_name:
            match_slug = slugify(f"{kickoff}-{home_name}-{away_name}")
        match = matches.get(match_slug or "")
        if match and source and source_match_id:
            ref_rows.append({
                "entity_type": "MATCH",
                "entity_id": match["match_id"],
                "source": source,
                "source_entity_type": "fixture",
                "source_entity_id": str(source_match_id),
                "source_entity_name": source_fixture_key or match.get("slug"),
                "confidence": 0.95,
                "is_primary": False,
                "payload": compact_payload(row),
            })
        for side, team_name_key, team_id_key in (
            ("HOME", "home_team_name", "home_team_id"),
            ("AWAY", "away_team_name", "away_team_id"),
        ):
            team_name = first_present(row, team_name_key)
            source_team_id = first_present(row, team_id_key)
            team = teams.get(team_key(team_name))
            if team and source and source_team_id:
                ref_rows.append({
                    "entity_type": "TEAM",
                    "entity_id": team["team_id"],
                    "source": source,
                    "source_entity_type": "team",
                    "source_entity_id": str(source_team_id),
                    "source_entity_name": str(team_name),
                    "confidence": 0.9,
                    "is_primary": False,
                    "payload": {"side": side, "source": SOURCE_NAME},
                })

    if ref_rows:
        sb.upsert("entity_external_refs", ref_rows, "entity_type,source,source_entity_id")


def compact_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if value not in (None, "")}


def upsert_match_detail_sheets(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    season_id: str,
    matches: dict[str, dict[str, Any]],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> None:
    match_indexes = build_match_indexes(data, matches)
    upsert_missing_players_from_detail_sheets(sb, data, teams, players)
    upsert_inferred_rosters_from_detail_sheets(sb, data, season_id, teams, players)
    upsert_player_match_stats_from_sheet(sb, data, match_indexes, teams, players)
    upsert_match_events_from_sheet(sb, data, match_indexes, teams, players)
    upsert_match_lineups_from_sheet(sb, data, match_indexes, teams, players)
    upsert_referees_from_sheet(sb, data, match_indexes)


def build_match_indexes(data: dict[str, list[dict[str, Any]]], matches: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    by_api_football = {}
    by_football_data = {}
    by_espn = {}
    for match in matches.values():
        lookup = match.get("_match_lookup") or {}
        for key, target in (
            ("api_football_fixture_id", by_api_football),
            ("football_data_match_id", by_football_data),
            ("espn_event_id", by_espn),
        ):
            value = source_id_text(lookup.get(key))
            if value:
                target[value] = match
    for row in get_rows(data, "EspnStats"):
        api_id = source_id_text(first_present(row, "fixture_id"))
        espn_id = source_id_text(first_present(row, "espn_event_id"))
        match = by_api_football.get(api_id or "")
        if match and espn_id:
            by_espn[espn_id] = match
    return {"API_FOOTBALL": by_api_football, "FOOTBALL_DATA": by_football_data, "ESPN": by_espn}


def match_from_fixture_id(match_indexes: dict[str, dict[str, Any]], fixture_id: Any) -> dict[str, Any] | None:
    raw = str(fixture_id or "").strip()
    if raw.startswith("espn_"):
        return match_indexes["ESPN"].get(raw.replace("espn_", "", 1))
    source_id = source_id_text(raw)
    return (
        match_indexes["API_FOOTBALL"].get(source_id or "")
        or match_indexes["FOOTBALL_DATA"].get(source_id or "")
        or match_indexes["ESPN"].get(source_id or "")
    )


def upsert_missing_players_from_detail_sheets(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> None:
    existing_slugs = set(players)
    rows = []
    candidates = []
    for sheet, id_key, name_key, team_key_name, source in (
        ("PlayerMatchStats", "player_id", "player_name", "team_name", "API_FOOTBALL"),
        ("EventosLive", "jugador_id", "jugador", "equipo", "API_FOOTBALL"),
        ("EventosLive", "assist_id", "assist", "equipo", "API_FOOTBALL"),
        ("ResumenJugadorPartido", "jugador_id", "jugador", "equipo", "API_FOOTBALL"),
        ("Alineaciones", "jugador_id", "jugador", "equipo", "ESPN"),
    ):
        for row in get_rows(data, sheet):
            source_id = source_id_text(first_present(row, id_key))
            name = first_present(row, name_key)
            team = teams.get(team_key(first_present(row, team_key_name)))
            if not source_id or not name:
                continue
            slug = f"player-{source.lower().replace('_', '-')}-{source_id}"
            if source == "API_FOOTBALL":
                slug = f"player-api-football-{source_id}"
            if slug in existing_slugs:
                continue
            existing_slugs.add(slug)
            rows.append({
                "slug": slug,
                "display_name": str(name).strip(),
                "normalized_name": normalize_text(name),
                "nationality_country_code": team["country_code"] if team else None,
                "gender": "MEN",
                "metadata": {"source": source, "inferred_from": sheet},
            })
            candidates.append((slug, source, source_id, name, team["team_id"] if team else None, row))
    saved = sb.upsert("players", rows, "slug", returning=True)
    saved_by_slug = {row["slug"]: row for row in saved}
    ref_rows = []
    membership_rows = []
    for slug, source, source_id, name, team_id, source_row in candidates:
        player = saved_by_slug.get(slug)
        if not player:
            continue
        player["_team_id"] = team_id
        player["_lookup_name"] = normalize_text(player["display_name"])
        players[slug] = player
        ref_rows.append(build_external_ref("PLAYER", player["player_id"], source, "player" if source == "API_FOOTBALL" else "athlete", source_id, name, compact_payload(source_row), 0.88))
        if team_id:
            membership_rows.append({
                "player_id": player["player_id"],
                "team_id": team_id,
                "membership_type": "NATIONAL_TEAM",
                "source": source,
                "confidence": 0.75,
                "metadata": {"inferred_from": "detail_sheets"},
            })
    sb.upsert("entity_external_refs", ref_rows, "entity_type,source,source_entity_id")
    sb.upsert("team_memberships", membership_rows, "player_id,team_id,membership_type,source")


def upsert_inferred_rosters_from_detail_sheets(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    season_id: str,
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> None:
    indexes = player_indexes(players)
    roster_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    membership_by_key: dict[tuple[str, str], dict[str, Any]] = {}

    def add_candidate(sheet: str, row: dict[str, Any], player_id_key: str, player_name_key: str, team_name_key: str) -> None:
        team = teams.get(team_key(first_present(row, team_name_key)))
        if not team:
            return
        player = match_player(indexes, first_present(row, player_id_key), first_present(row, player_name_key), team["team_id"])
        if not player:
            return
        position = first_present(row, "position", "posicion", "player_position", "pos")
        roster_by_key[(season_id, team["team_id"], player["player_id"])] = {
            "competition_season_id": season_id,
            "team_id": team["team_id"],
            "player_id": player["player_id"],
            "shirt_number": to_int(first_present(row, "number", "numero", "shirt_number")),
            "position": str(position).strip() if position not in (None, "") else None,
            "roster_status": "UNKNOWN",
            "metadata": {"source": SOURCE_NAME, "inferred_from": sheet},
        }
        membership_by_key[(player["player_id"], team["team_id"])] = {
            "player_id": player["player_id"],
            "team_id": team["team_id"],
            "membership_type": "NATIONAL_TEAM",
            "source": f"{SOURCE_NAME}:INFERRED_ROSTER",
            "confidence": 0.7,
            "metadata": {"inferred_from": sheet},
        }

    for row in get_rows(data, "PlayerMatchStats"):
        add_candidate("PlayerMatchStats", row, "player_id", "player_name", "team_name")
    for row in get_rows(data, "ResumenJugadorPartido"):
        add_candidate("ResumenJugadorPartido", row, "jugador_id", "jugador", "equipo")
    for row in get_rows(data, "EventosLive"):
        add_candidate("EventosLive", row, "jugador_id", "jugador", "equipo")
        add_candidate("EventosLive", row, "assist_id", "assist", "equipo")
    for row in get_rows(data, "Alineaciones"):
        add_candidate("Alineaciones", row, "jugador_id", "jugador", "equipo")

    sb.upsert("team_memberships", list(membership_by_key.values()), "player_id,team_id,membership_type,source")
    sb.upsert("competition_rosters", list(roster_by_key.values()), "competition_season_id,team_id,player_id")


def upsert_player_match_stats_from_sheet(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    match_indexes: dict[str, dict[str, Any]],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> None:
    indexes = player_indexes(players)
    stat_rows = []
    stat_fields = {
        "minutes_played": "minutes_played",
        "rating": "rating",
        "shots_total": "shots_total",
        "shots_on": "shots_on",
        "goals_scored": "goals_scored",
        "goals_conceded": "goals_conceded",
        "assists": "assists",
        "passes_total": "passes_total",
        "passes_accuracy": "passes_accuracy",
        "key_passes": "key_passes",
        "tackles_total": "tackles_total",
        "interceptions": "interceptions",
        "blocks": "blocks",
        "duels_total": "duels_total",
        "duels_won": "duels_won",
        "dribbles_attempts": "dribbles_attempts",
        "dribbles_success": "dribbles_success",
        "fouls_committed": "fouls_committed",
        "fouls_drawn": "fouls_drawn",
        "yellow_cards": "yellow_cards",
        "red_cards": "red_cards",
    }
    for row in get_rows(data, "PlayerMatchStats"):
        match = match_from_fixture_id(match_indexes, first_present(row, "fixture_id"))
        team = teams.get(team_key(first_present(row, "team_name")))
        player = match_player(indexes, first_present(row, "player_id"), first_present(row, "player_name"), team["team_id"] if team else None)
        if not match or not team or not player:
            continue
        for source_field, stat_name in stat_fields.items():
            value = first_present(row, source_field)
            if value in (None, ""):
                continue
            stat_rows.append({
                "match_id": match["match_id"],
                "team_id": team["team_id"],
                "player_id": player["player_id"],
                "stat_name": stat_name,
                "stat_value": value,
                "source": "API_FOOTBALL",
                "captured_at": parse_datetime({"date_utc": first_present(row, "loaded_at")}, "UTC") or datetime.now(timezone.utc).isoformat(),
                "payload": compact_payload(row),
            })
    sb.upsert("player_match_stats", stat_rows, "match_id,player_id,stat_name,source")


def upsert_match_events_from_sheet(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    match_indexes: dict[str, dict[str, Any]],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> None:
    indexes = player_indexes(players)
    event_rows = []
    for row in get_rows(data, "EventosLive"):
        match = match_from_fixture_id(match_indexes, first_present(row, "fixture_id", "match_id"))
        team = teams.get(team_key(first_present(row, "equipo")))
        player = match_player(indexes, first_present(row, "jugador_id"), first_present(row, "jugador"), team["team_id"] if team else None)
        assist = match_player(indexes, first_present(row, "assist_id"), first_present(row, "assist"), team["team_id"] if team else None)
        if not match:
            continue
        event_rows.append({
            "match_id": match["match_id"],
            "team_id": team["team_id"] if team else None,
            "player_id": player["player_id"] if player else None,
            "related_player_id": assist["player_id"] if assist else None,
            "event_type": str(first_present(row, "tipo_evento") or "UNKNOWN"),
            "event_detail": first_present(row, "detalle_evento"),
            "minute": to_int(first_present(row, "minuto")),
            "stoppage_minute": to_int(first_present(row, "extra")),
            "occurred_at": None,
            "source": "API_FOOTBALL",
            "source_event_id": str(first_present(row, "evento_id") or f"{first_present(row, 'fixture_id')}_{first_present(row, 'minuto')}_{first_present(row, 'tipo_evento')}_{first_present(row, 'jugador_id')}"),
            "payload": compact_payload(row),
        })
    sb.upsert("match_events", event_rows, "source,source_event_id")


def upsert_match_lineups_from_sheet(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    match_indexes: dict[str, dict[str, Any]],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> None:
    indexes = player_indexes(players)
    rows = []
    for row in get_rows(data, "Alineaciones"):
        match = match_from_fixture_id(match_indexes, first_present(row, "fixture_id"))
        team = teams.get(team_key(first_present(row, "equipo")))
        player = match_player(indexes, first_present(row, "jugador_id"), first_present(row, "jugador"), team["team_id"] if team else None)
        if not match or not team or not player:
            continue
        role = normalize_text(first_present(row, "rol"))
        rows.append({
            "match_id": match["match_id"],
            "team_id": team["team_id"],
            "player_id": player["player_id"],
            "lineup_role": "STARTER" if role in {"titular", "starter"} else "SUBSTITUTE" if role in {"suplente", "substitute"} else "UNKNOWN",
            "position": first_present(row, "posicion"),
            "shirt_number": to_int(first_present(row, "numero")),
            "is_captain": False,
            "source": "ESPN" if str(first_present(row, "fixture_id") or "").startswith("espn_") else SOURCE_NAME,
            "metadata": compact_payload(row),
        })
    sb.upsert("match_lineups", rows, "match_id,team_id,player_id,source")


def upsert_referees_from_sheet(sb: Supabase, data: dict[str, list[dict[str, Any]]], match_indexes: dict[str, dict[str, Any]]) -> None:
    referee_rows = []
    official_specs = []
    for row in get_rows(data, "Arbitros"):
        name = first_present(row, "nombre")
        referee_source_id = source_id_text(first_present(row, "arbitro_id")) or slugify(name)
        if not name:
            continue
        slug = f"referee-{slugify(name)}"
        referee_rows.append({
            "slug": slug,
            "display_name": str(name).strip(),
            "normalized_name": normalize_text(name),
            "nationality_country_code": None,
            "metadata": compact_payload(row),
        })
        match = match_from_fixture_id(match_indexes, first_present(row, "fixture_id"))
        if match:
            official_specs.append((slug, match["match_id"], referee_source_id, row))
    saved = sb.upsert("referees", referee_rows, "slug", returning=True)
    referees = {row["slug"]: row for row in saved}
    official_rows = []
    ref_rows = []
    for slug, match_id, source_id, source_row in official_specs:
        referee = referees.get(slug)
        if not referee:
            continue
        official_rows.append({
            "match_id": match_id,
            "referee_id": referee["referee_id"],
            "role": "REFEREE",
            "metadata": compact_payload(source_row),
        })
        ref_rows.append(build_external_ref("REFEREE", referee["referee_id"], SOURCE_NAME, "referee", source_id, referee["display_name"], compact_payload(source_row), 0.8))
    sb.upsert("match_officials", official_rows, "match_id,referee_id,role")
    sb.upsert("entity_external_refs", ref_rows, "entity_type,source,source_entity_id")


def enrich_from_optional_apis(
    sb: Supabase,
    args: argparse.Namespace,
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
    matches: dict[str, dict[str, Any]],
    venues: dict[str, dict[str, Any]],
) -> None:
    api_refs: list[dict[str, Any]] = []

    api_football_key = os.environ.get("API_FOOTBALL_KEY")
    if args.api_football_budget > 0 and api_football_key:
        client = BudgetedApiClient(
            "API_FOOTBALL",
            "https://v3.football.api-sports.io",
            {"x-apisports-key": api_football_key},
            budget=args.api_football_budget,
            cache_dir=args.api_cache_dir,
            verify_tls=not args.insecure_skip_tls_verify,
        )
        for payload in fetch_api_football_today_window(client, args.api_football_league, args.api_football_season, args.api_football_window_days):
            api_refs.extend(refs_from_api_football_fixture(payload, teams, matches))
            api_refs.extend(player_refs_from_api_football_payload(payload, players, teams))
    elif args.api_football_budget > 0:
        print("Skipping API_FOOTBALL enrichment: set API_FOOTBALL_KEY.")

    football_data_key = os.environ.get("FOOTBALL_DATA_TOKEN")
    if args.football_data_budget > 0 and football_data_key:
        client = BudgetedApiClient(
            "FOOTBALL_DATA",
            "https://api.football-data.org/v4",
            {"X-Auth-Token": football_data_key},
            budget=args.football_data_budget,
            cache_dir=args.api_cache_dir,
            verify_tls=not args.insecure_skip_tls_verify,
        )
        for payload in fetch_football_data_world_cup(
            client,
            args.football_data_competition_code,
            args.api_football_season,
            args.football_data_date_from,
            args.football_data_date_to,
        ):
            api_refs.extend(refs_from_football_data(payload, teams, matches))
            api_refs.extend(player_refs_from_football_data_payload(payload, players, teams))
            apply_football_data_match_updates(sb, payload, matches)
    elif args.football_data_budget > 0:
        print("Skipping FOOTBALL_DATA enrichment: set FOOTBALL_DATA_TOKEN.")

    if args.espn_budget > 0:
        client = BudgetedApiClient(
            "ESPN",
            "https://site.api.espn.com/apis/site/v2/sports/soccer",
            {},
            budget=args.espn_budget,
            cache_dir=args.api_cache_dir,
            verify_tls=not args.insecure_skip_tls_verify,
        )
        for payload in fetch_espn_scoreboard_window(client, args.espn_league, args.api_football_window_days):
            api_refs.extend(refs_from_espn_scoreboard(payload, teams, matches))
            api_refs.extend(player_refs_from_espn_payload(payload, players, teams))

    sportmonks_token = os.environ.get("SPORTMONKS_API_TOKEN")
    if args.sportmonks_budget > 0 and sportmonks_token:
        client = BudgetedApiClient(
            "SPORTMONKS",
            "https://api.sportmonks.com/v3",
            {},
            budget=args.sportmonks_budget,
            cache_dir=args.api_cache_dir,
            verify_tls=not args.insecure_skip_tls_verify,
            fail_on_error=False,
        )
        country_payloads: list[dict[str, Any]] = []
        if not args.sportmonks_skip_countries:
            country_payloads = fetch_sportmonks_countries(client, sportmonks_token, args.sportmonks_country_pages)
            api_refs.extend(apply_sportmonks_country_refs(sb, country_payloads, teams))
        if country_payloads and not args.sportmonks_skip_country_players:
            for payload in fetch_sportmonks_players_by_country(client, sportmonks_token, country_payloads, teams):
                api_refs.extend(player_refs_from_sportmonks_payload(payload, players))
        for payload in fetch_sportmonks_fixtures(
            client,
            sportmonks_token,
            args.sportmonks_fixture_date_from,
            args.sportmonks_fixture_date_to,
            args.sportmonks_fixture_include,
            args.sportmonks_fixture_pages,
        ):
            api_refs.extend(apply_sportmonks_fixture_payload(sb, payload, teams, players, matches))
    elif args.sportmonks_budget > 0:
        print("Skipping SPORTMONKS enrichment: set SPORTMONKS_API_TOKEN.")

    if api_refs:
        sb.upsert("entity_external_refs", api_refs, "entity_type,source,source_entity_id")
        print(json.dumps({"external_refs_from_optional_apis": len(api_refs)}, ensure_ascii=False))


def fetch_api_football_today_window(client: BudgetedApiClient, league: int, season: int, days: int) -> list[dict[str, Any]]:
    out = []
    today = datetime.now(timezone.utc).date()
    for offset in range(max(1, days)):
        day = today.fromordinal(today.toordinal() - offset)
        payload = client.get("fixtures", {"league": league, "season": season, "date": day.isoformat()})
        if isinstance(payload, dict):
            out.extend(payload.get("response") or [])
    return out


def fetch_football_data_world_cup(
    client: BudgetedApiClient,
    competition_code: str,
    season: int,
    date_from: str | None,
    date_to: str | None,
) -> list[dict[str, Any]]:
    out = []
    match_query: dict[str, Any] = {"season": season}
    if date_from:
        match_query["dateFrom"] = date_from
    if date_to:
        match_query["dateTo"] = date_to
    matches = client.get(f"competitions/{competition_code}/matches", match_query)
    standings = client.get(f"competitions/{competition_code}/standings", None)
    if isinstance(matches, dict):
        out.append({"resource": "matches", "payload": matches})
    if isinstance(standings, dict):
        out.append({"resource": "standings", "payload": standings})
    return out


def fetch_espn_scoreboard_window(client: BudgetedApiClient, league: str, days: int) -> list[dict[str, Any]]:
    out = []
    today = datetime.now(timezone.utc).date()
    for offset in range(max(1, days)):
        day = today.fromordinal(today.toordinal() - offset)
        payload = client.get(f"{league}/scoreboard", {"dates": day.strftime("%Y%m%d")})
        if isinstance(payload, dict):
            out.append(payload)
    return out


def fetch_sportmonks_countries(client: BudgetedApiClient, token: str, max_pages: int) -> list[dict[str, Any]]:
    out = []
    for page in range(1, max(1, max_pages) + 1):
        payload = client.get("core/countries", {"api_token": token, "per_page": 50, "page": page})
        if not isinstance(payload, dict):
            break
        out.append(payload)
        pagination = payload.get("pagination") or {}
        has_more = pagination.get("has_more")
        if has_more is False:
            break
        if not (payload.get("data") or []):
            break
    return out


def fetch_sportmonks_players_by_country(
    client: BudgetedApiClient,
    token: str,
    country_payloads: list[dict[str, Any]],
    teams: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    countries_by_alpha2 = sportmonks_countries_by_alpha2(country_payloads)
    out = []
    requested_country_ids = set()
    for team in teams.values():
        country_code = team.get("country_code")
        country = countries_by_alpha2.get(str(country_code or "").upper())
        country_id = source_id_text(country.get("id") if country else None)
        if not country_id or country_id in requested_country_ids:
            continue
        requested_country_ids.add(country_id)
        payload = client.get(f"football/players/countries/{country_id}", {"api_token": token})
        if isinstance(payload, dict):
            out.append({"country": country, "payload": payload})
    return out


def fetch_sportmonks_fixtures(
    client: BudgetedApiClient,
    token: str,
    date_from: str,
    date_to: str,
    include: str,
    max_pages: int,
) -> list[dict[str, Any]]:
    out = []
    for page in range(1, max(1, max_pages) + 1):
        payload = client.get(
            f"football/fixtures/between/{date_from}/{date_to}",
            {"api_token": token, "include": include, "per_page": 50, "page": page},
        )
        if not isinstance(payload, dict):
            break
        out.append({"resource": "fixtures", "payload": payload, "page": page})
        pagination = payload.get("pagination") or {}
        has_more = pagination.get("has_more")
        if has_more is False:
            break
        if not (payload.get("data") or []):
            break
    return out


def sportmonks_countries_by_alpha2(country_payloads: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out = {}
    for payload in country_payloads:
        for country in payload.get("data") or []:
            if not isinstance(country, dict):
                continue
            alpha2 = first_present(country, "iso2", "code", "country_code", "code_alpha2")
            if alpha2:
                out[str(alpha2).upper()] = country
    return out


def apply_sportmonks_country_refs(
    sb: Supabase,
    country_payloads: list[dict[str, Any]],
    teams: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    countries_by_alpha2 = sportmonks_countries_by_alpha2(country_payloads)
    if not countries_by_alpha2:
        return []

    refs = []
    existing_countries = sb.select("countries", {"select": "code_alpha2,payload"})
    for country in existing_countries:
        alpha2 = str(country.get("code_alpha2") or "").upper()
        sportmonks_country = countries_by_alpha2.get(alpha2)
        sportmonks_id = source_id_text(sportmonks_country.get("id") if sportmonks_country else None)
        if not sportmonks_id:
            continue
        payload = dict(country.get("payload") or {})
        external_refs = dict(payload.get("external_refs") or {})
        external_refs["SPORTMONKS"] = {
            "source_entity_type": "country",
            "source_entity_id": sportmonks_id,
            "source_entity_name": first_present(sportmonks_country, "name", "official_name"),
            "payload": compact_payload(sportmonks_country),
        }
        payload["external_refs"] = external_refs
        sb.update("countries", {"payload": payload}, {"code_alpha2": f"eq.{alpha2}"})

    for team in teams.values():
        sportmonks_country = countries_by_alpha2.get(str(team.get("country_code") or "").upper())
        sportmonks_id = source_id_text(sportmonks_country.get("id") if sportmonks_country else None)
        if sportmonks_id:
            refs.append(build_external_ref("TEAM", team["team_id"], "SPORTMONKS", "country", sportmonks_id, sportmonks_country.get("name"), sportmonks_country, 0.6))
    return refs


def apply_sportmonks_fixture_payload(
    sb: Supabase,
    payload: dict[str, Any],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
    matches: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    body = payload.get("payload") or {}
    refs: list[dict[str, Any]] = []
    for fixture in body.get("data") or []:
        if not isinstance(fixture, dict):
            continue
        participants = sportmonks_participants(fixture)
        home_payload = participants.get("HOME")
        away_payload = participants.get("AWAY")
        home_name = sportmonks_participant_name(home_payload)
        away_name = sportmonks_participant_name(away_payload)
        matched = find_match_by_api_names(matches, home_name, away_name, first_present(fixture, "starting_at", "starting_at_timestamp"))
        if not matched:
            continue
        fixture_id = source_id_text(fixture.get("id"))
        if fixture_id:
            refs.append(build_external_ref("MATCH", matched["match_id"], "SPORTMONKS", "fixture", fixture_id, sportmonks_fixture_name(home_name, away_name), fixture, 0.8))
        for side, participant in participants.items():
            team = teams.get(team_key(sportmonks_participant_name(participant)))
            source_id = source_id_text((participant or {}).get("id"))
            if team and source_id:
                refs.append(build_external_ref("TEAM", team["team_id"], "SPORTMONKS", "participant", source_id, sportmonks_participant_name(participant), participant, 0.75))

        home_score, away_score = sportmonks_scores(fixture)
        match_update = {
            "metadata": (matched.get("metadata") or {}) | {"sportmonks": compact_payload(fixture)},
        }
        status = sportmonks_match_status(fixture)
        if status:
            match_update["status"] = status
        if home_score is not None and away_score is not None:
            match_update["home_score"] = home_score
            match_update["away_score"] = away_score
        sb.update("matches", match_update, {"match_id": f"eq.{matched['match_id']}"})
        if home_score is not None:
            sb.update("match_participants", {"score": home_score}, {"match_id": f"eq.{matched['match_id']}", "side": "eq.HOME"})
        if away_score is not None:
            sb.update("match_participants", {"score": away_score}, {"match_id": f"eq.{matched['match_id']}", "side": "eq.AWAY"})

        refs.extend(upsert_sportmonks_lineups(sb, fixture, matched, teams, players))
        refs.extend(upsert_sportmonks_events(sb, fixture, matched, teams, players))
        refs.extend(upsert_sportmonks_statistics(sb, fixture, matched, teams, players))
    return refs


def sportmonks_fixture_name(home: Any, away: Any) -> str:
    return f"{home or ''} vs {away or ''}".strip()


def sportmonks_participants(fixture: dict[str, Any]) -> dict[str, dict[str, Any]]:
    raw = fixture.get("participants") or []
    if isinstance(raw, dict):
        raw = raw.get("data") or raw.get("data", [])
    out: dict[str, dict[str, Any]] = {}
    for participant in raw if isinstance(raw, list) else []:
        if not isinstance(participant, dict):
            continue
        meta = participant.get("meta") or {}
        location = str(first_present(meta, "location") or first_present(participant, "location") or "").upper()
        if location in {"HOME", "AWAY"}:
            out[location] = participant
    return out


def sportmonks_participant_name(participant: dict[str, Any] | None) -> Any:
    if not participant:
        return None
    return first_present(participant, "name", "display_name", "short_code")


def sportmonks_scores(fixture: dict[str, Any]) -> tuple[int | None, int | None]:
    scores = fixture.get("scores") or []
    if isinstance(scores, dict):
        scores = scores.get("data") or []
    home = away = None
    for score in scores if isinstance(scores, list) else []:
        if not isinstance(score, dict):
            continue
        description = normalize_text(first_present(score, "description", "type", "score_type"))
        if description and not any(token in description for token in ("current", "fulltime", "full time", "2nd half")):
            continue
        participant = str(first_present(score, "score_participant", "participant") or "").lower()
        goals = to_int(first_present(score, "score", "goals", "value"))
        if goals is None and isinstance(score.get("score"), dict):
            goals = to_int(first_present(score["score"], "goals", "participant"))
        if participant == "home":
            home = goals
        elif participant == "away":
            away = goals
    return home, away


def sportmonks_match_status(fixture: dict[str, Any]) -> str | None:
    state = fixture.get("state") or {}
    raw = str(first_present(state, "name", "short_name", "developer_name") or first_present(fixture, "state_id") or "").upper()
    if raw in {"FT", "AET", "FT_PEN", "FINISHED", "FULL_TIME"}:
        return "FINISHED"
    if raw in {"LIVE", "HT", "INPLAY", "1ST_HALF", "2ND_HALF"}:
        return "LIVE"
    if raw in {"POSTPONED", "SUSPENDED"}:
        return "POSTPONED"
    if raw in {"CANCELLED", "CANCELED"}:
        return "CANCELLED"
    return None


def upsert_sportmonks_lineups(
    sb: Supabase,
    fixture: dict[str, Any],
    matched: dict[str, Any],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    lineups = nested_list(fixture.get("lineups"))
    indexes = player_indexes(players)
    rows = []
    refs = []
    for item in lineups:
        team = sportmonks_team_from_payload(item, teams)
        player_payload = item.get("player") if isinstance(item.get("player"), dict) else item
        source_id = source_id_text(first_present(item, "player_id") or player_payload.get("id"))
        name = first_present(player_payload, "display_name", "common_name", "name", "fullname", "full_name")
        player = match_player(indexes, source_id, name, team["team_id"] if team else None)
        if not team or not player:
            continue
        if source_id:
            refs.append(build_external_ref("PLAYER", player["player_id"], "SPORTMONKS", "player", source_id, name, player_payload, 0.8))
        rows.append({
            "match_id": matched["match_id"],
            "team_id": team["team_id"],
            "player_id": player["player_id"],
            "lineup_role": sportmonks_lineup_role(item),
            "position": first_present(item, "position", "position_name", "formation_position"),
            "shirt_number": to_int(first_present(item, "jersey_number", "number", "shirt_number")),
            "is_captain": bool(first_present(item, "captain", "is_captain") or False),
            "source": "SPORTMONKS",
            "metadata": compact_payload(item),
        })
    sb.upsert("match_lineups", rows, "match_id,team_id,player_id,source")
    return refs


def upsert_sportmonks_events(
    sb: Supabase,
    fixture: dict[str, Any],
    matched: dict[str, Any],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    events = nested_list(fixture.get("events"))
    indexes = player_indexes(players)
    rows = []
    refs = []
    for event in events:
        event_id = source_id_text(event.get("id"))
        if not event_id:
            continue
        team = sportmonks_team_from_payload(event, teams)
        player_payload = event.get("player") if isinstance(event.get("player"), dict) else {}
        related_payload = event.get("related_player") if isinstance(event.get("related_player"), dict) else {}
        player_id = source_id_text(first_present(event, "player_id") or player_payload.get("id"))
        related_id = source_id_text(first_present(event, "related_player_id") or related_payload.get("id"))
        player = match_player(indexes, player_id, sportmonks_player_name(player_payload), team["team_id"] if team else None)
        related = match_player(indexes, related_id, sportmonks_player_name(related_payload), team["team_id"] if team else None)
        if player and player_id:
            refs.append(build_external_ref("PLAYER", player["player_id"], "SPORTMONKS", "player", player_id, sportmonks_player_name(player_payload), player_payload, 0.8))
        if related and related_id:
            refs.append(build_external_ref("PLAYER", related["player_id"], "SPORTMONKS", "player", related_id, sportmonks_player_name(related_payload), related_payload, 0.75))
        rows.append({
            "match_id": matched["match_id"],
            "team_id": team["team_id"] if team else None,
            "player_id": player["player_id"] if player else None,
            "related_player_id": related["player_id"] if related else None,
            "event_type": str(first_present(event, "type", "type_name", "event_type") or first_present(event.get("type") or {}, "name", "code") or "UNKNOWN"),
            "event_detail": first_present(event, "info", "addition", "result"),
            "minute": to_int(first_present(event, "minute")),
            "stoppage_minute": to_int(first_present(event, "extra_minute", "injury_time")),
            "source": "SPORTMONKS",
            "source_event_id": event_id,
            "payload": compact_payload(event),
        })
    sb.upsert("match_events", rows, "source,source_event_id")
    return refs


def upsert_sportmonks_statistics(
    sb: Supabase,
    fixture: dict[str, Any],
    matched: dict[str, Any],
    teams: dict[str, dict[str, Any]],
    players: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    stats = nested_list(fixture.get("statistics"))
    indexes = player_indexes(players)
    rows = []
    refs = []
    for stat in stats:
        player_payload = stat.get("player") if isinstance(stat.get("player"), dict) else {}
        source_id = source_id_text(first_present(stat, "player_id") or player_payload.get("id"))
        if not source_id and not player_payload:
            continue
        team = sportmonks_team_from_payload(stat, teams)
        player = match_player(indexes, source_id, sportmonks_player_name(player_payload), team["team_id"] if team else None)
        if not team or not player:
            continue
        if source_id:
            refs.append(build_external_ref("PLAYER", player["player_id"], "SPORTMONKS", "player", source_id, sportmonks_player_name(player_payload), player_payload, 0.8))
        for stat_name, stat_value in sportmonks_stat_values(stat).items():
            rows.append({
                "match_id": matched["match_id"],
                "team_id": team["team_id"],
                "player_id": player["player_id"],
                "stat_name": stat_name,
                "stat_value": stat_value,
                "source": "SPORTMONKS",
                "captured_at": datetime.now(timezone.utc).isoformat(),
                "payload": compact_payload(stat),
            })
    sb.upsert("player_match_stats", rows, "match_id,player_id,stat_name,source")
    return refs


def nested_list(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, dict):
        value = value.get("data") or value.get("items") or []
    return [item for item in value if isinstance(item, dict)] if isinstance(value, list) else []


def sportmonks_team_from_payload(payload: dict[str, Any], teams: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    participant = payload.get("participant") if isinstance(payload.get("participant"), dict) else payload.get("team") if isinstance(payload.get("team"), dict) else {}
    name = first_present(participant, "name", "display_name") or first_present(payload, "team_name", "participant_name")
    return teams.get(team_key(name))


def sportmonks_player_name(payload: dict[str, Any]) -> Any:
    return first_present(payload, "display_name", "common_name", "name", "fullname", "full_name")


def sportmonks_lineup_role(payload: dict[str, Any]) -> str:
    raw = normalize_text(first_present(payload, "type", "lineup_type", "formation_position"))
    if any(token in raw for token in ("starting", "starter", "startxi", "lineup")):
        return "STARTER"
    if any(token in raw for token in ("bench", "substitute")):
        return "SUBSTITUTE"
    return "UNKNOWN"


def sportmonks_stat_values(payload: dict[str, Any]) -> dict[str, float]:
    candidates = payload.get("details") or payload.get("values") or payload.get("statistics") or []
    out: dict[str, float] = {}
    if isinstance(candidates, dict):
        candidates = candidates.get("data") or list(candidates.values())
    for item in candidates if isinstance(candidates, list) else []:
        if not isinstance(item, dict):
            continue
        raw_name = first_present(item, "type", "type_name", "name", "code")
        value = first_present(item, "value", "data", "amount")
        stat_value = to_float(value)
        stat_name = sportmonks_stat_name(raw_name)
        if stat_name and stat_value is not None:
            out[stat_name] = stat_value
    for raw_name in ("rating", "minutes", "minutes_played", "goals", "assists", "yellow_cards", "red_cards"):
        stat_value = to_float(payload.get(raw_name))
        stat_name = sportmonks_stat_name(raw_name)
        if stat_name and stat_value is not None:
            out[stat_name] = stat_value
    return out


def sportmonks_stat_name(value: Any) -> str | None:
    raw = normalize_text(value)
    mapping = {
        "minutes": "minutes_played",
        "minutes played": "minutes_played",
        "rating": "rating",
        "goals": "goals_scored",
        "goal": "goals_scored",
        "assists": "assists",
        "assist": "assists",
        "yellow cards": "yellow_cards",
        "yellow card": "yellow_cards",
        "red cards": "red_cards",
        "red card": "red_cards",
        "shots total": "shots_total",
        "shots on target": "shots_on",
        "passes": "passes_total",
        "key passes": "key_passes",
        "tackles": "tackles_total",
        "interceptions": "interceptions",
    }
    return mapping.get(raw) or (raw.replace(" ", "_") if raw else None)


def refs_from_api_football_fixture(payload: dict[str, Any], teams: dict[str, dict[str, Any]], matches: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    refs = []
    fixture = payload.get("fixture") or {}
    teams_payload = payload.get("teams") or {}
    fixture_id = fixture.get("id")
    if fixture_id:
        matched = find_match_by_api_names(matches, teams_payload.get("home", {}).get("name"), teams_payload.get("away", {}).get("name"), fixture.get("date"))
        if matched:
            refs.append(build_external_ref("MATCH", matched["match_id"], "API_FOOTBALL", "fixture", fixture_id, fixture.get("referee") or matched.get("slug"), payload, 0.85))
    for side in ("home", "away"):
        item = teams_payload.get(side) or {}
        team = teams.get(team_key(item.get("name")))
        if team and item.get("id"):
            refs.append(build_external_ref("TEAM", team["team_id"], "API_FOOTBALL", "team", item["id"], item.get("name"), item, 0.85))
    return refs


def refs_from_football_data(payload: dict[str, Any], teams: dict[str, dict[str, Any]], matches: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    refs = []
    body = payload.get("payload") or {}
    if payload.get("resource") == "matches":
        for match_payload in body.get("matches") or []:
            home = (match_payload.get("homeTeam") or {}).get("name")
            away = (match_payload.get("awayTeam") or {}).get("name")
            matched = find_match_by_api_names(matches, home, away, match_payload.get("utcDate"))
            if matched and match_payload.get("id"):
                refs.append(build_external_ref("MATCH", matched["match_id"], "FOOTBALL_DATA", "match", match_payload["id"], f"{home} vs {away}", match_payload, 0.85))
            for team_key_name in ("homeTeam", "awayTeam"):
                item = match_payload.get(team_key_name) or {}
                team = teams.get(team_key(item.get("name")))
                if team and item.get("id"):
                    refs.append(build_external_ref("TEAM", team["team_id"], "FOOTBALL_DATA", "team", item["id"], item.get("name"), item, 0.85))
    return refs


def apply_football_data_match_updates(sb: Supabase, payload: dict[str, Any], matches: dict[str, dict[str, Any]]) -> None:
    if payload.get("resource") != "matches":
        return
    body = payload.get("payload") or {}
    for match_payload in body.get("matches") or []:
        home = (match_payload.get("homeTeam") or {}).get("name")
        away = (match_payload.get("awayTeam") or {}).get("name")
        matched = find_match_by_api_names(matches, home, away, match_payload.get("utcDate"))
        if not matched:
            continue
        home_score, away_score = football_data_full_time_score(match_payload)
        status = football_data_match_status(match_payload.get("status"))
        match_update = {
            "status": status,
            "metadata": (matched.get("metadata") or {}) | {"football_data": compact_payload(match_payload)},
        }
        if home_score is not None and away_score is not None:
            match_update["home_score"] = home_score
            match_update["away_score"] = away_score
        sb.update("matches", match_update, {"match_id": f"eq.{matched['match_id']}"})
        if home_score is not None:
            sb.update("match_participants", {"score": home_score}, {"match_id": f"eq.{matched['match_id']}", "side": "eq.HOME"})
        if away_score is not None:
            sb.update("match_participants", {"score": away_score}, {"match_id": f"eq.{matched['match_id']}", "side": "eq.AWAY"})


def football_data_full_time_score(match_payload: dict[str, Any]) -> tuple[int | None, int | None]:
    score = match_payload.get("score") or {}
    full_time = score.get("fullTime") or {}
    return to_int(full_time.get("home")), to_int(full_time.get("away"))


def football_data_match_status(value: Any) -> str:
    raw = str(value or "").upper()
    if raw in {"FINISHED"}:
        return "FINISHED"
    if raw in {"IN_PLAY", "PAUSED"}:
        return "LIVE"
    if raw in {"POSTPONED", "SUSPENDED"}:
        return "POSTPONED"
    if raw in {"CANCELLED"}:
        return "CANCELLED"
    return "SCHEDULED"


def refs_from_espn_scoreboard(payload: dict[str, Any], teams: dict[str, dict[str, Any]], matches: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    refs = []
    for event in payload.get("events") or []:
        competitions = event.get("competitions") or []
        competitors = (competitions[0].get("competitors") if competitions else []) or []
        home = next((c for c in competitors if c.get("homeAway") == "home"), {})
        away = next((c for c in competitors if c.get("homeAway") == "away"), {})
        home_name = (home.get("team") or {}).get("displayName")
        away_name = (away.get("team") or {}).get("displayName")
        matched = find_match_by_api_names(matches, home_name, away_name, event.get("date"))
        if matched and event.get("id"):
            refs.append(build_external_ref("MATCH", matched["match_id"], "ESPN", "event", event["id"], event.get("name"), event, 0.8))
        for competitor in (home, away):
            item = competitor.get("team") or {}
            team = teams.get(team_key(item.get("displayName") or item.get("shortDisplayName")))
            if team and item.get("id"):
                refs.append(build_external_ref("TEAM", team["team_id"], "ESPN", "team", item["id"], item.get("displayName"), item, 0.8))
    return refs


def player_refs_from_api_football_payload(payload: dict[str, Any], players: dict[str, dict[str, Any]], teams: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    indexes = player_indexes(players)
    refs = []
    for item in iter_api_football_player_items(payload):
        player_payload = item["player"]
        source_id = player_payload.get("id")
        name = player_payload.get("name")
        team = teams.get(team_key(item.get("team_name")))
        player = match_player(indexes, source_id, name, team["team_id"] if team else None)
        if player and source_id:
            refs.append(build_external_ref("PLAYER", player["player_id"], "API_FOOTBALL", "player", source_id, name, player_payload, 0.95 if str(source_id) in indexes["api_football_id"] else 0.82))
    return refs


def player_refs_from_football_data_payload(payload: dict[str, Any], players: dict[str, dict[str, Any]], teams: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    indexes = player_indexes(players)
    refs = []
    for item in iter_football_data_player_items(payload.get("payload") or {}):
        person = item["player"]
        source_id = person.get("id")
        name = person.get("name")
        team = teams.get(team_key(item.get("team_name")))
        player = match_player(indexes, source_id, name, team["team_id"] if team else None)
        if player and source_id:
            refs.append(build_external_ref("PLAYER", player["player_id"], "FOOTBALL_DATA", "person", source_id, name, person, 0.82))
    return refs


def player_refs_from_espn_payload(payload: dict[str, Any], players: dict[str, dict[str, Any]], teams: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    indexes = player_indexes(players)
    refs = []
    for item in iter_espn_athlete_items(payload):
        athlete = item["player"]
        source_id = athlete.get("id")
        name = athlete.get("displayName") or athlete.get("fullName") or athlete.get("shortName")
        team = teams.get(team_key(item.get("team_name")))
        player = match_player(indexes, source_id, name, team["team_id"] if team else None)
        if player and source_id:
            refs.append(build_external_ref("PLAYER", player["player_id"], "ESPN", "athlete", source_id, name, athlete, 0.78))
    return refs


def player_refs_from_sportmonks_payload(payload: dict[str, Any], players: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    by_unique_name = unique_player_name_index(players)
    refs = []
    body = payload.get("payload") or {}
    for player_payload in iter_sportmonks_player_items(body):
        source_id = player_payload.get("id")
        name = first_present(player_payload, "display_name", "common_name", "name", "fullname", "full_name")
        player = by_unique_name.get(normalize_text(name))
        if player and source_id:
            refs.append(build_external_ref("PLAYER", player["player_id"], "SPORTMONKS", "player", source_id, name, player_payload, 0.72))
    return refs


def player_indexes(players: dict[str, dict[str, Any]]) -> dict[str, Any]:
    by_api_football_id = {}
    by_espn_id = {}
    by_name_team = {}
    for player in players.values():
        slug = str(player.get("slug") or "")
        if slug.startswith("player-api-football-"):
            by_api_football_id[slug.replace("player-api-football-", "", 1)] = player
        if slug.startswith("player-espn-"):
            by_espn_id[slug.replace("player-espn-", "", 1)] = player
        team_id = player.get("_team_id")
        name = player.get("_lookup_name") or normalize_text(player.get("display_name"))
        if team_id and name:
            by_name_team[(name, team_id)] = player
    return {"api_football_id": by_api_football_id, "espn_id": by_espn_id, "name_team": by_name_team}


def unique_player_name_index(players: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    buckets: dict[str, list[dict[str, Any]]] = {}
    for player in players.values():
        name = normalize_text(player.get("_lookup_name") or player.get("display_name"))
        if name:
            buckets.setdefault(name, []).append(player)
    return {name: rows[0] for name, rows in buckets.items() if len(rows) == 1}


def match_player(indexes: dict[str, Any], source_id: Any, name: Any, team_id: Any) -> dict[str, Any] | None:
    normalized_source_id = source_id_text(source_id)
    if normalized_source_id and normalized_source_id in indexes["api_football_id"]:
        return indexes["api_football_id"][normalized_source_id]
    if normalized_source_id and normalized_source_id in indexes["espn_id"]:
        return indexes["espn_id"][normalized_source_id]
    if name and team_id:
        return indexes["name_team"].get((normalize_text(name), team_id))
    return None


def iter_api_football_player_items(payload: dict[str, Any]) -> Iterable[dict[str, Any]]:
    for lineup in payload.get("lineups") or []:
        team_name = (lineup.get("team") or {}).get("name")
        for bucket in ("startXI", "substitutes"):
            for entry in lineup.get(bucket) or []:
                player = entry.get("player") or {}
                if player.get("id") or player.get("name"):
                    yield {"team_name": team_name, "player": player}
    for event in payload.get("events") or []:
        team_name = (event.get("team") or {}).get("name")
        for key in ("player", "assist"):
            player = event.get(key) or {}
            if player.get("id") or player.get("name"):
                yield {"team_name": team_name, "player": player}
    for stat in payload.get("players") or []:
        team_name = (stat.get("team") or {}).get("name")
        for entry in stat.get("players") or []:
            player = entry.get("player") or {}
            if player.get("id") or player.get("name"):
                yield {"team_name": team_name, "player": player}


def iter_football_data_player_items(payload: dict[str, Any]) -> Iterable[dict[str, Any]]:
    for match_payload in payload.get("matches") or []:
        for team_key_name in ("homeTeam", "awayTeam"):
            team_name = (match_payload.get(team_key_name) or {}).get("name")
            for key in ("scorers", "lineup", "bench"):
                for person in match_payload.get(key) or []:
                    if isinstance(person, dict) and (person.get("id") or person.get("name")):
                        yield {"team_name": team_name, "player": person}
    for person in payload.get("squad") or []:
        if isinstance(person, dict) and (person.get("id") or person.get("name")):
            yield {"team_name": None, "player": person}


def iter_espn_athlete_items(payload: dict[str, Any]) -> Iterable[dict[str, Any]]:
    for event in payload.get("events") or []:
        competitions = event.get("competitions") or []
        competitors = (competitions[0].get("competitors") if competitions else []) or []
        for competitor in competitors:
            team_name = (competitor.get("team") or {}).get("displayName")
            for key in ("leaders", "statistics", "roster"):
                for bucket in competitor.get(key) or []:
                    if isinstance(bucket, dict):
                        athletes = bucket.get("athletes") or bucket.get("leaders") or []
                        for item in athletes:
                            athlete = item.get("athlete") if isinstance(item, dict) else None
                            if athlete and (athlete.get("id") or athlete.get("displayName") or athlete.get("fullName")):
                                yield {"team_name": team_name, "player": athlete}


def iter_sportmonks_player_items(payload: dict[str, Any]) -> Iterable[dict[str, Any]]:
    for item in payload.get("data") or []:
        if isinstance(item, dict) and (item.get("id") or first_present(item, "display_name", "common_name", "name", "fullname", "full_name")):
            yield item


def build_external_ref(entity_type: str, entity_id: str, source: str, source_entity_type: str, source_entity_id: Any, source_entity_name: Any, payload: dict[str, Any], confidence: float) -> dict[str, Any]:
    return {
        "entity_type": entity_type,
        "entity_id": entity_id,
        "source": source,
        "source_entity_type": source_entity_type,
        "source_entity_id": str(source_entity_id),
        "source_entity_name": str(source_entity_name or ""),
        "confidence": confidence,
        "is_primary": False,
        "payload": payload,
    }


def find_match_by_api_names(matches: dict[str, dict[str, Any]], home_name: Any, away_name: Any, kickoff: Any) -> dict[str, Any] | None:
    if not home_name or not away_name or not kickoff:
        return None
    kickoff_iso = parse_datetime({"date_utc": kickoff}, "UTC") or str(kickoff)
    kickoff_date = kickoff_iso[:10]
    wanted_home = team_key(home_name)
    wanted_away = team_key(away_name)
    for match in matches.values():
        lookup = match.get("_match_lookup") or {}
        if (
            lookup.get("kickoff_date") == kickoff_date
            and lookup.get("home_team_key") == wanted_home
            and lookup.get("away_team_key") == wanted_away
        ):
            return match
    for match in matches.values():
        lookup = match.get("_match_lookup") or {}
        if (
            lookup.get("kickoff_date") == kickoff_date
            and {lookup.get("home_team_key"), lookup.get("away_team_key")} == {wanted_home, wanted_away}
        ):
            return match
    return None


def normalize_country_from_text(value: Any) -> str | None:
    if not value:
        return None
    text = normalize_text(value).replace(" ", "")
    return {
        "usa": "US",
        "unitedstates": "US",
        "estadosunidos": "US",
        "mexico": "MX",
        "canada": "CA",
    }.get(text)


def upsert_matches(
    sb: Supabase,
    data: dict[str, list[dict[str, Any]]],
    season_id: str,
    stages: dict[str, dict[str, Any]],
    teams: dict[str, dict[str, Any]],
    venues: dict[str, dict[str, Any]],
    source_tz: str,
) -> dict[str, dict[str, Any]]:
    group_rows = sb.select("competition_groups", {"select": "*", "competition_season_id": f"eq.{season_id}"})
    groups = {row["group_code"]: row for row in group_rows}

    match_rows = []
    participant_specs = []
    match_lookup: dict[str, dict[str, Any]] = {}
    slot_rows_by_code: dict[str, dict[str, Any]] = {}
    for row in get_rows(data, "Partidos"):
        home_raw = first_present(row, "local", "home_team_name")
        away_raw = first_present(row, "visitante", "away_team_name")
        kickoff = parse_datetime(row, source_tz)
        if not home_raw or not away_raw or not kickoff:
            continue
        home_key = team_key(home_raw) if not is_tournament_slot(home_raw) else None
        away_key = team_key(away_raw) if not is_tournament_slot(away_raw) else None
        home_group_code = team_group_code(home_key, data) if home_key else None
        away_group_code = team_group_code(away_key, data) if away_key else None
        stage_code = stage_code_from_row(row)
        stage_code = stage_code_from_participants(row, home_raw, away_raw, stage_code, source_tz)
        group = groups.get(str(first_present(row, "grupo", "group_name") or "").strip())
        if home_key and away_key and home_group_code and home_group_code == away_group_code:
            stage_code = "GROUP_STAGE"
            group = groups.get(home_group_code) or group
        venue_slug = canonical_venue_slug(first_present(row, "estadio", "venue_name")) if first_present(row, "estadio", "venue_name") else None
        slug = slugify(first_present(row, "match_id", "match_key") or f"{kickoff}-{home_raw}-{away_raw}")
        match_lookup[slug] = {
            "home_team_key": team_key(home_raw) if not is_tournament_slot(home_raw) else None,
            "away_team_key": team_key(away_raw) if not is_tournament_slot(away_raw) else None,
            "kickoff_date": kickoff[:10],
            "home_raw": str(home_raw),
            "away_raw": str(away_raw),
            "api_football_fixture_id": source_id_text(first_present(row, "fixture_id_api_football", "fixture_id")),
            "football_data_match_id": source_id_text(first_present(row, "match_id_football_data")),
            "source_match_id": source_id_text(first_present(row, "match_id")),
        }
        match_rows.append({
            "competition_season_id": season_id,
            "stage_id": stages[stage_code]["stage_id"],
            "group_id": group["group_id"] if group else None,
            "venue_id": venues.get(venue_slug, {}).get("venue_id") if venue_slug else None,
            "slug": slug,
            "match_number": to_int(first_present(row, "match_number", "matchday")),
            "kickoff_at": kickoff,
            "status": match_status(first_present(row, "status")),
            "is_neutral": True,
            "home_score": to_int(first_present(row, "goles_local", "home_score")),
            "away_score": to_int(first_present(row, "goles_visitante", "away_score")),
            "metadata": {"source": SOURCE_NAME},
        })
        participant_specs.append((slug, "HOME", home_raw, to_int(first_present(row, "goles_local", "home_score")), stage_code, group))
        participant_specs.append((slug, "AWAY", away_raw, to_int(first_present(row, "goles_visitante", "away_score")), stage_code, group))
        for raw in (home_raw, away_raw):
            if is_tournament_slot(raw):
                code = slot_code(raw)
                slot_rows_by_code[code] = {
                    "competition_season_id": season_id,
                    "stage_id": stages[stage_code]["stage_id"],
                    "slot_code": code,
                    "slot_label": str(raw).strip(),
                    "slot_type": infer_slot_type(raw),
                    "source_group_id": group["group_id"] if group else None,
                    "metadata": {"source": SOURCE_NAME},
                }

    saved_matches = sb.upsert("matches", match_rows, "slug", returning=True)
    matches = {row["slug"]: row for row in saved_matches}
    for slug, lookup in match_lookup.items():
        if slug in matches:
            matches[slug]["_match_lookup"] = lookup
    saved_slots = sb.upsert("tournament_slots", list(slot_rows_by_code.values()), "competition_season_id,slot_code", returning=True)
    slots = {row["slot_code"]: row for row in saved_slots}
    participant_rows = []
    for match_slug, side, raw, score, _stage_code, _group in participant_specs:
        match = matches.get(match_slug)
        if not match:
            continue
        base_participant = {
            "match_id": match["match_id"],
            "side": side,
            "team_id": None,
            "tournament_slot_id": None,
            "is_home_designation": side == "HOME",
            "score": score,
            "metadata": {"source": SOURCE_NAME, "raw_name": raw},
        }
        if is_tournament_slot(raw):
            slot = slots.get(slot_code(raw))
            if not slot:
                continue
            participant_rows.append(base_participant | {
                "participant_role": "SLOT",
                "tournament_slot_id": slot["tournament_slot_id"],
            })
        else:
            team = teams.get(team_key(raw))
            if not team:
                continue
            participant_rows.append(base_participant | {
                "participant_role": "TEAM",
                "team_id": team["team_id"],
            })
    sb.upsert("match_participants", participant_rows, "match_id,side")
    return matches


def stage_code_from_row(row: dict[str, Any]) -> str:
    raw = normalize_text(first_present(row, "fase", "stage", "round") or "")
    if "group" in raw or "grupo" in raw:
        return "GROUP_STAGE"
    if "32" in raw or "dieciseis" in raw:
        return "ROUND_OF_32"
    if "16" in raw or "octav" in raw:
        return "ROUND_OF_16"
    if "quarter" in raw or "cuarto" in raw:
        return "QUARTER_FINAL"
    if "semi" in raw:
        return "SEMI_FINAL"
    if "third" in raw or "tercer" in raw:
        return "THIRD_PLACE"
    if "final" in raw:
        return "FINAL"
    return "GROUP_STAGE"


def stage_code_from_participants(
    row: dict[str, Any],
    home_raw: Any,
    away_raw: Any,
    fallback: str,
    source_tz: str,
) -> str:
    if fallback != "GROUP_STAGE":
        return fallback
    combined = normalize_text(f"{home_raw or ''} {away_raw or ''}")
    if not (is_tournament_slot(home_raw) or is_tournament_slot(away_raw)):
        return fallback
    if "semifinal" in combined and "loser" in combined:
        return "THIRD_PLACE"
    if "semifinal" in combined and "winner" in combined:
        return "FINAL"
    if "quarter" in combined and "winner" in combined:
        return "SEMI_FINAL"
    if ("round of 16" in combined or "octav" in combined) and "winner" in combined:
        return "QUARTER_FINAL"
    if ("round of 32" in combined or "dieciseis" in combined) and "winner" in combined:
        return "ROUND_OF_16"
    if "group" in combined or "grupo" in combined or "third place group" in combined:
        return "ROUND_OF_32"

    kickoff = parse_datetime(row, source_tz)
    if kickoff:
        date_key = kickoff[:10]
        if "2026-06-28" <= date_key <= "2026-07-03":
            return "ROUND_OF_32"
        if "2026-07-04" <= date_key <= "2026-07-07":
            return "ROUND_OF_16"
        if "2026-07-09" <= date_key <= "2026-07-11":
            return "QUARTER_FINAL"
        if "2026-07-14" <= date_key <= "2026-07-15":
            return "SEMI_FINAL"
        if date_key == "2026-07-18":
            return "THIRD_PLACE"
        if date_key == "2026-07-19":
            return "FINAL"
    return fallback


def match_status(value: Any) -> str:
    raw = normalize_text(value)
    if raw in {"ft", "finished", "finalizado"}:
        return "FINISHED"
    if raw in {"live", "inplay", "ht"}:
        return "LIVE"
    if raw in {"postponed"}:
        return "POSTPONED"
    if raw in {"cancelled", "canceled"}:
        return "CANCELLED"
    return "SCHEDULED"


def infer_slot_type(value: Any) -> str:
    raw = normalize_text(value)
    if "third place group" in raw:
        return "BEST_THIRD"
    if "winner" in raw:
        return "WINNER"
    if "loser" in raw:
        return "LOSER"
    if "2nd" in raw or "second" in raw:
        return "GROUP_RUNNER_UP"
    return "STRUCTURAL"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Migrate WC2026 source data to clean Supabase schema.")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--xlsx", help="Path to workbook export.")
    source.add_argument("--google-spreadsheet-id", help="Live source spreadsheet ID.")
    parser.add_argument("--google-credentials-json", default=os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"), help="Service account JSON with readonly access to the source spreadsheet.")
    parser.add_argument("--supabase-url", default=os.environ.get("SUPABASE_URL"))
    parser.add_argument("--supabase-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_KEY"))
    parser.add_argument("--season-slug", default=DEFAULT_SEASON_SLUG)
    parser.add_argument("--competition-slug", default=DEFAULT_COMPETITION_SLUG)
    parser.add_argument("--source-timezone", default=DEFAULT_SOURCE_TZ)
    parser.add_argument("--venues-file", default=os.environ.get("VENUES_FILE"), help="TSV with Location Name, Latitude and Longitude.")
    parser.add_argument("--api-cache-dir", default=".cache/api_enrichment", help="Local cache for optional API enrichment responses.")
    parser.add_argument("--api-football-budget", type=int, default=0, help="Max API-Football requests for this run. Default 0 avoids spending free quota.")
    parser.add_argument("--api-football-league", type=int, default=1)
    parser.add_argument("--api-football-season", type=int, default=2026)
    parser.add_argument("--api-football-window-days", type=int, default=2, help="Fetch today/yesterday window when API-Football is enabled.")
    parser.add_argument("--football-data-budget", type=int, default=0, help="Max football-data.org requests for this run. Default 0.")
    parser.add_argument("--football-data-competition-code", default="WC")
    parser.add_argument("--football-data-date-from", default="2026-06-11", help="Start date for football-data.org match enrichment.")
    parser.add_argument("--football-data-date-to", default=date.today().isoformat(), help="End date for football-data.org match enrichment.")
    parser.add_argument("--espn-budget", type=int, default=0, help="Max ESPN scoreboard requests for this run. Default 0.")
    parser.add_argument("--espn-league", default="fifa.world")
    parser.add_argument("--sportmonks-budget", type=int, default=0, help="Max Sportmonks requests for this run. Default 0.")
    parser.add_argument("--sportmonks-country-pages", type=int, default=6, help="Max Sportmonks country catalog pages to scan before player enrichment.")
    parser.add_argument("--sportmonks-skip-countries", action="store_true", help="Skip Sportmonks country catalog enrichment and preserve budget for fixtures.")
    parser.add_argument("--sportmonks-skip-country-players", action="store_true", help="Skip Sportmonks players-by-country enrichment and preserve budget for fixtures.")
    parser.add_argument("--sportmonks-fixture-date-from", default="2026-06-11", help="Start date for Sportmonks fixture enrichment.")
    parser.add_argument("--sportmonks-fixture-date-to", default=date.today().isoformat(), help="End date for Sportmonks fixture enrichment.")
    parser.add_argument("--sportmonks-fixture-pages", type=int, default=20, help="Max paginated Sportmonks fixture pages to fetch.")
    parser.add_argument("--sportmonks-fixture-include", default="participants;scores;state;venue;events;lineups;statistics;referees;weatherReport", help="Sportmonks fixture includes.")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sleep-seconds", type=float, default=0.0)
    parser.add_argument("--insecure-skip-tls-verify", action="store_true", help="Disable TLS verification for Supabase requests. Use only as a temporary local workaround.")
    return parser.parse_args()


def main() -> None:
    global SOURCE_NAME
    args = parse_args()
    if not args.supabase_url or not args.supabase_key:
        raise SystemExit("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY/SUPABASE_KEY.")
    if args.google_spreadsheet_id:
        SOURCE_NAME = "LIVE_SOURCE_SPREADSHEET"
        data = load_google_spreadsheet(args.google_spreadsheet_id, args.google_credentials_json)
    else:
        SOURCE_NAME = "WORKBOOK_EXPORT"
        data = load_xlsx(args.xlsx)
    sb = Supabase(args.supabase_url, args.supabase_key, args.dry_run, args.sleep_seconds, not args.insecure_skip_tls_verify)
    migrate(data, sb, args)


if __name__ == "__main__":
    main()
